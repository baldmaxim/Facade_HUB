"""Post-run evaluator for priced VOR workbooks.

This module inspects the produced Excel output and surfaces suspicious
positions. It is intentionally independent from the main runtime path so it can
be used as an external review loop during iterative development.
"""

from __future__ import annotations

import argparse
import json
import re
from dataclasses import asdict, dataclass, field
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook
import yaml


COMPONENT_TYPES = {"Р", "М", "МХ", "ТР"}


@dataclass
class EvaluationIssue:
    severity: str
    code: str
    message: str


@dataclass
class EvaluatedComponent:
    row_index: int
    row_type: str
    name: str
    unit: str
    quantity: float | None
    unit_price: float | None
    total: float | None


@dataclass
class EvaluatedPosition:
    row_index: int
    title: str
    unit: str
    quantity: float | None
    customer_note: str
    gp_note: str
    components: list[EvaluatedComponent] = field(default_factory=list)
    issues: list[EvaluationIssue] = field(default_factory=list)
    inferred_domain: str = "general"
    computed_total: float = 0.0
    computed_per_unit: float = 0.0


@dataclass
class EvaluationReport:
    workbook: str
    total_positions: int
    issue_positions: int
    error_positions: int
    warning_positions: int
    positions: list[EvaluatedPosition]

    def to_markdown(self, max_positions: int = 25) -> str:
        lines = [
            f"# VOR Evaluation Report: {self.workbook}",
            "",
            f"- Positions: {self.total_positions}",
            f"- Positions with issues: {self.issue_positions}",
            f"- Error positions: {self.error_positions}",
            f"- Warning positions: {self.warning_positions}",
            "",
        ]

        ranked = sorted(
            self.positions,
            key=lambda p: (
                -sum(1 for i in p.issues if i.severity == "ERROR"),
                -sum(1 for i in p.issues if i.severity == "WARNING"),
                p.row_index,
            ),
        )

        shown = 0
        for pos in ranked:
            if not pos.issues:
                continue
            shown += 1
            if shown > max_positions:
                break
            lines.append(
                f"## Row {pos.row_index}: {pos.title}"
            )
            lines.append(f"- Domain: {pos.inferred_domain}")
            if pos.quantity:
                lines.append(f"- Quantity: {pos.quantity:g} {pos.unit}")
            if pos.computed_per_unit > 0:
                lines.append(f"- Computed per-unit total: {pos.computed_per_unit:,.2f}")
            lines.append(f"- Components: {len(pos.components)}")
            for issue in pos.issues:
                lines.append(
                    f"- [{issue.severity}] `{issue.code}`: {issue.message}"
                )
            lines.append("")

        return "\n".join(lines).rstrip() + "\n"


class WorkbookEvaluator:
    """Evaluate a priced VOR workbook using workbook-level heuristics."""

    def __init__(self, price_ranges_path: str | Path | None = None) -> None:
        if price_ranges_path is None:
            price_ranges_path = Path(__file__).with_name("price_ranges.yaml")
        with open(price_ranges_path, "r", encoding="utf-8") as fh:
            self._ranges = yaml.safe_load(fh) or {}

    def evaluate_workbook(
        self,
        workbook_path: str | Path,
        *,
        focus_domains: set[str] | None = None,
    ) -> EvaluationReport:
        workbook_path = Path(workbook_path)
        wb = load_workbook(workbook_path, read_only=True, data_only=True)
        try:
            ws = wb[wb.sheetnames[0]]
            positions = self._extract_positions(ws.iter_rows(values_only=True))
        finally:
            wb.close()

        for position in positions:
            self._evaluate_position(position)

        if focus_domains:
            positions = [
                p for p in positions if p.inferred_domain in focus_domains
            ]

        issue_positions = sum(1 for p in positions if p.issues)
        error_positions = sum(
            1 for p in positions if any(i.severity == "ERROR" for i in p.issues)
        )
        warning_positions = sum(
            1 for p in positions if any(i.severity == "WARNING" for i in p.issues)
        )

        return EvaluationReport(
            workbook=workbook_path.name,
            total_positions=len(positions),
            issue_positions=issue_positions,
            error_positions=error_positions,
            warning_positions=warning_positions,
            positions=positions,
        )

    def _extract_positions(
        self,
        rows: Iterable[tuple],
    ) -> list[EvaluatedPosition]:
        positions: list[EvaluatedPosition] = []
        current: EvaluatedPosition | None = None

        for row_index, row in enumerate(rows, start=1):
            if row_index == 1:
                continue

            marker = self._str(row, 2)
            name = self._str(row, 6)
            unit = self._str(row, 7)
            quantity = self._float(row, 8)
            unit_price = self._float(row, 15)
            total = self._float(row, 16)

            if marker in COMPONENT_TYPES:
                if current is None:
                    continue
                current.components.append(
                    EvaluatedComponent(
                        row_index=row_index,
                        row_type=marker,
                        name=name,
                        unit=unit,
                        quantity=quantity,
                        unit_price=unit_price,
                        total=total,
                    )
                )
                continue

            if not any(
                [
                    self._str(row, 1),
                    self._str(row, 2),
                    name,
                    unit,
                    quantity is not None,
                ]
            ):
                continue

            current = EvaluatedPosition(
                row_index=row_index,
                title=name,
                unit=unit,
                quantity=quantity,
                customer_note=self._str(row, 18),
                gp_note=self._str(row, 19),
            )
            positions.append(current)

        return positions

    def _evaluate_position(self, position: EvaluatedPosition) -> None:
        position.inferred_domain = self._infer_domain(position.title)
        position.computed_total = round(
            sum(
                c.total
                if c.total is not None
                else (c.quantity or 0.0) * (c.unit_price or 0.0)
                for c in position.components
            ),
            2,
        )
        if position.quantity and position.quantity > 0:
            position.computed_per_unit = round(
                position.computed_total / position.quantity, 2
            )

        self._check_suspicious_notes(position)
        self._check_component_presence(position)
        self._check_zero_prices(position)
        self._check_benchmarks(position)
        self._check_component_depth(position)

    def _check_suspicious_notes(self, position: EvaluatedPosition) -> None:
        note = " ".join(
            part for part in [position.customer_note, position.gp_note] if part
        ).lower()
        suspicious = [
            ("fallback_note", "стандартная расценка"),
            ("fallback_note", "применительно"),
            ("candidate_failure", "кандидат"),
            ("missing_source", "не найдена"),
            ("missing_price", "no fer price found"),
            ("missing_price", "цена не найдена"),
        ]
        for code, needle in suspicious:
            if needle in note:
                position.issues.append(
                    EvaluationIssue(
                        severity="WARNING",
                        code=code,
                        message=f"Suspicious note fragment detected: '{needle}'",
                    )
                )
        match = re.search(r"уверенность:\s*(\d+)%", note)
        if match and int(match.group(1)) < 60:
            position.issues.append(
                EvaluationIssue(
                    severity="WARNING",
                    code="weak_confidence",
                    message=f"Low confidence note detected: {match.group(1)}%",
                )
            )

    def _check_component_presence(self, position: EvaluatedPosition) -> None:
        names = [c.name.lower() for c in position.components]
        row_types = {c.row_type for c in position.components}

        if "Р" not in row_types:
            position.issues.append(
                EvaluationIssue(
                    severity="ERROR",
                    code="missing_work_row",
                    message="Position has no work row ('Р') in the output.",
                )
            )

        if position.inferred_domain == "concrete":
            self._require_any(position, names, ["бетон"], "missing_concrete")
            self._require_any(position, names, ["арматур"], "missing_rebar")
            self._recommend_any(
                position,
                names,
                ["опалуб", "насос", "кран", "вибратор"],
                "thin_concrete_composition",
                "Concrete composition looks too thin for estimator-grade output.",
            )

        if position.inferred_domain == "masonry":
            title = position.title.lower()
            if "газобет" in title or "блок" in title:
                self._require_any(position, names, ["блок"], "missing_blocks")
                self._require_any(position, names, ["кле"], "missing_glue")
                self._recommend_any(
                    position,
                    names,
                    ["раствор", "сетк", "арм", "анкер", "u-блок", "u блок"],
                    "thin_masonry_composition",
                    "Masonry composition lacks common auxiliary/support items.",
                )
            elif "кирпич" in title:
                self._require_any(position, names, ["кирпич"], "missing_brick")
                self._require_any(position, names, ["раствор"], "missing_mortar")

    def _check_zero_prices(self, position: EvaluatedPosition) -> None:
        for component in position.components:
            if component.unit_price is None:
                continue
            if component.unit_price > 0:
                continue
            if component.row_type not in COMPONENT_TYPES:
                continue
            position.issues.append(
                EvaluationIssue(
                    severity="ERROR",
                    code="zero_price_component",
                    message=(
                        f"{component.row_type} row '{component.name}' has zero price."
                    ),
                )
            )

    def _check_benchmarks(self, position: EvaluatedPosition) -> None:
        if position.computed_per_unit <= 0:
            return

        position_ranges = self._ranges.get("position_totals", {})
        title = position.title.lower()
        unit = position.unit.lower()

        for bench in position_ranges.values():
            keywords = [str(k).lower() for k in bench.get("keywords", [])]
            if keywords and not any(k in title for k in keywords):
                continue

            expected_unit = str(bench.get("unit", "")).lower()
            if expected_unit and expected_unit != unit:
                continue

            min_value = float(bench.get("min", 0))
            max_value = float(bench.get("max", 0))
            if min_value and position.computed_per_unit < min_value:
                position.issues.append(
                    EvaluationIssue(
                        severity="WARNING",
                        code="benchmark_low",
                        message=(
                            f"Per-unit total {position.computed_per_unit:,.2f} is below "
                            f"benchmark {min_value:,.2f}-{max_value:,.2f}."
                        ),
                    )
                )
            if max_value and position.computed_per_unit > max_value:
                position.issues.append(
                    EvaluationIssue(
                        severity="ERROR",
                        code="benchmark_high",
                        message=(
                            f"Per-unit total {position.computed_per_unit:,.2f} is above "
                            f"benchmark {min_value:,.2f}-{max_value:,.2f}."
                        ),
                    )
                )
            return

    def _check_component_depth(self, position: EvaluatedPosition) -> None:
        if position.inferred_domain == "concrete" and len(position.components) < 5:
            position.issues.append(
                EvaluationIssue(
                    severity="WARNING",
                    code="too_few_components",
                    message="Concrete position has fewer than 5 output lines.",
                )
            )
        if position.inferred_domain == "masonry" and len(position.components) < 4:
            position.issues.append(
                EvaluationIssue(
                    severity="WARNING",
                    code="too_few_components",
                    message="Masonry position has fewer than 4 output lines.",
                )
            )

    def _infer_domain(self, title: str) -> str:
        lower = title.lower()
        if any(word in lower for word in ["кладк", "газобет", "кирпич", "перегородк", "вентшахт"]):
            return "masonry"
        if any(word in lower for word in ["монолит", "бетон", "ж/б", "железобет", "опалуб", "арматур"]):
            return "concrete"
        return "general"

    def _require_any(
        self,
        position: EvaluatedPosition,
        names: list[str],
        needles: list[str],
        code: str,
    ) -> None:
        if any(any(needle in name for needle in needles) for name in names):
            return
        position.issues.append(
            EvaluationIssue(
                severity="ERROR",
                code=code,
                message=f"Missing required component matching: {', '.join(needles)}",
            )
        )

    def _recommend_any(
        self,
        position: EvaluatedPosition,
        names: list[str],
        needles: list[str],
        code: str,
        message: str,
    ) -> None:
        if any(any(needle in name for needle in needles) for name in names):
            return
        position.issues.append(
            EvaluationIssue(
                severity="WARNING",
                code=code,
                message=message,
            )
        )

    @staticmethod
    def _str(row: tuple, col_1_based: int) -> str:
        value = row[col_1_based - 1] if len(row) >= col_1_based else None
        return str(value).strip() if value is not None else ""

    @staticmethod
    def _float(row: tuple, col_1_based: int) -> float | None:
        value = row[col_1_based - 1] if len(row) >= col_1_based else None
        if value is None:
            return None
        if isinstance(value, (int, float)):
            return float(value)
        try:
            return float(str(value).replace(" ", "").replace(",", "."))
        except ValueError:
            return None


def _build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Evaluate a priced VOR workbook.")
    parser.add_argument("workbook", help="Path to priced .xlsx workbook")
    parser.add_argument(
        "--json",
        action="store_true",
        help="Print JSON instead of Markdown summary",
    )
    parser.add_argument(
        "--max-positions",
        type=int,
        default=25,
        help="Maximum number of problematic positions to show in Markdown mode",
    )
    parser.add_argument(
        "--focus",
        nargs="*",
        default=None,
        help="Limit evaluation to specific inferred domains, e.g. concrete masonry",
    )
    return parser


def main() -> int:
    parser = _build_arg_parser()
    args = parser.parse_args()

    evaluator = WorkbookEvaluator()
    focus_domains = set(args.focus) if args.focus else None
    report = evaluator.evaluate_workbook(
        args.workbook,
        focus_domains=focus_domains,
    )

    if args.json:
        print(json.dumps(asdict(report), ensure_ascii=False, indent=2))
    else:
        print(report.to_markdown(max_positions=args.max_positions))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
