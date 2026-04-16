"""Run VOR V5 pipeline using file-based LLM (Claude Code as LLM).

Phase 1: Extract batches → vor_batches/
Phase 3: Read responses from vor_responses/ → price → assemble Excel

Usage:
    # Phase 1: extract batches
    python vor_claude_runner.py extract

    # Phase 3: after Claude generates responses
    python vor_claude_runner.py assemble
"""
import asyncio
import json
import logging
import os
import sys
import time
from pathlib import Path

backend_dir = Path(__file__).resolve().parent
sys.path.insert(0, str(backend_dir))
os.chdir(str(backend_dir))

from dotenv import load_dotenv
load_dotenv(backend_dir / ".env")

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s %(name)s: %(message)s",
)
logger = logging.getLogger("vor_claude")

VOR_PATH = Path("Расчет ПЗ_ЖК Сокольники_Версия 1.xlsx")
BATCHES_DIR = Path("vor_batches")
RESPONSES_DIR = Path("vor_responses")
OUTPUT_PATH = Path("vor_priced_v5_claude.xlsx")
GESN_DB = str(backend_dir / "data" / "gesn.db")
PROJECT_ROOT = "."


def _stem_russian(word: str) -> str:
    """Very rough Russian stemmer: strip common suffixes to get word root.

    Handles plurals, case endings, and diminutives for construction terms.
    E.g.: бульдозеры -> бульдозер, экскаваторы -> экскаватор,
          смеси -> смес, краны -> кран, катки -> катк.
    """
    w = word.lower().rstrip(",.")
    # Longest suffixes first
    for suffix in (
        "ами", "ями", "ями", "ов", "ев", "ей", "ий", "ых", "их",
        "ые", "ие", "ой", "ый", "ая", "яя", "ую", "юю",
        "ы", "и", "а", "я", "о", "е", "у", "ю",
    ):
        if len(w) > len(suffix) + 3 and w.endswith(suffix):
            return w[: -len(suffix)]
    return w


# Machinery alias map: VOR composition name keywords → DB name keywords
# The DB uses formal ФССЦ names which differ from common VOR names
_MACHINERY_ALIASES = {
    "бетононасос": ["автобетононасос", "бетононасос"],
    "вибратор": ["вибратор", "вибробулав"],
    "подъёмник": ["подъемник", "автогидроподъемник", "мачтовый подъемник"],
    "подъемник": ["подъемник", "автогидроподъемник"],
    "леса": ["подмости", "леса"],
    "автомобиль": ["автомобил"],
    "автосамосвал": ["автомобил", "самосвал"],
    "экскаватор": ["экскаватор"],
    "бульдозер": ["бульдозер"],
    "растворосмеситель": ["растворосмесител"],
    "краскопульт": ["краскопульт", "распылител"],
    "каток": ["катк"],
    "автокран": ["кран", "автомобильн"],
    "кран": ["кран"],
    "автобетоносмеситель": ["автобетоносмесител", "бетоносмесител"],
    "виброрейка": ["виброрейк"],
    "кран/лебёдка": ["кран", "лебедк", "лебёдк"],
    "лебёдка": ["лебедк", "лебёдк"],
    "компрессор": ["компрессор"],
    "сварочный": ["сварочн", "электросварочн"],
}

# Material alias map: common VOR names → DB search keywords
_MATERIAL_ALIASES = {
    "бетон": ["смеси бетонн", "бетон"],
    "раствор": ["раствор", "смеси растворн"],
    "опалубка": ["опалубк", "щит опалубк"],
    "арматура": ["арматур", "сталь арматурн"],
    "кирпич": ["кирпич"],
    "газобетон": ["газобетон", "газосиликат", "ячеист"],
    "утеплитель": ["утеплител", "теплоизоляц", "минплит", "пенополистирол"],
    "гидроизоляция": ["гидроизоляц", "мастик"],
    "грунтовка": ["грунтовк"],
    "шпатлёвка": ["шпатлевк", "шпаклевк"],
    "шпатлевка": ["шпатлевк", "шпаклевк"],
    "клей": ["клей", "клеев"],
    "монтажная": ["пена монтажн", "пена полиуретан"],
    "пена": ["пена"],
    "сетка": ["сетк"],
    "профиль": ["профил"],
    "гипсокартон": ["гипсокартон", "листы гипсокартон"],
    "штукатурка": ["штукатурк", "смеси штукатурн"],
    "керамогранит": ["керамогранит", "плитк керамогранит"],
    "плитка": ["плитк"],
    "краска": ["краск", "эмал"],
    "стеклопакет": ["стеклопакет"],
    "трубы": ["труб"],
    "труба": ["труб"],
    "кабель": ["кабел"],
    "провод": ["провод"],
    "щебень": ["щебен", "щебн"],
    "песок": ["песок", "песч"],
    "цемент": ["цемент"],
    "доска": ["доск", "пиломатериал"],
    "брус": ["брус", "пиломатериал"],
    "гвозди": ["гвозд"],
    "болты": ["болт"],
    "анкер": ["анкер"],
    "лист": ["лист"],
    "блок": ["блок"],
    # Extended aliases for unpriced materials
    "монтажная пена": ["пена полиуретан", "пена монтажн"],
    "ГСМ": ["топливо дизель", "бензин"],
    "крепёж": ["метиз", "гвозд", "болт", "шуруп", "саморез", "дюбел"],
    "крепеж": ["метиз", "гвозд", "болт", "шуруп", "саморез", "дюбел"],
    "метизы": ["метиз", "гвозд", "болт", "шуруп", "саморез"],
    "ограждение": ["ограждени", "забор"],
    "рулонная": ["рулонн", "гидроизоляц", "изоляц"],
    "праймер": ["праймер", "грунтовк"],
    "минераловатный": ["минераловат", "минплит", "теплоизоляц"],
    "пенополистирол": ["пенополистирол", "экструзионн"],
    "плёнка": ["пленк", "плёнк"],
    "пленка": ["пленк", "плёнк"],
    "мембрана": ["мембран", "гидроизоляц", "пароизоляц"],
    "скотч": ["лента", "скотч"],
    "маяки": ["маяк", "маячков"],
    "наливной": ["наливн", "самовыравн", "стяжк"],
    "стяжка": ["стяжк", "раствор"],
    "звукоизоляционная": ["звукоизоляц", "шумоизоляц"],
    "электроды": ["электрод"],
    "металлоконструкции": ["металлоконструкц", "сталь конструкц"],
    "материалы": ["материал"],
    "оборудование": ["оборудовани"],
    "трубопроводы": ["трубопровод", "труб"],
    "кабельная": ["кабел", "провод"],
    "плиты": ["плит"],
    "баннер": ["баннер", "информационн"],
    "штора": ["штор"],
}

# Typical unit prices (2025 rubles) for fallback pricing by material category + unit
# Based on median prices from priced items in this project
_TYPICAL_PRICES: dict[tuple[str, str], float] = {
    # Materials (М)
    ("М", "кг"): 224,           # generic per-kg material
    ("М", "л"): 150,            # generic per-liter (primer, paint)
    ("М", "м.п."): 161,         # generic per-meter-run
    ("М", "м2"): 1500,          # generic per-m2 material (conservative)
    ("М", "м3"): 5634,          # generic per-m3 material
    ("М", "т"): 75660,          # generic per-tonne
    ("М", "шт"): 500,           # generic per-piece (conservative)
    ("М", "компл"): 8549,       # generic per-set
    ("М", "компл."): 8549,      # generic per-set (with dot)
    # Work (Р)
    ("Р", "м2"): 1801,          # generic work per-m2
    ("Р", "м3"): 1371,          # generic work per-m3
    ("Р", "м.п."): 1903,        # generic work per-meter-run
    ("Р", "шт"): 1801,          # generic work per-piece
    ("Р", "тн"): 15000,         # generic work per-tonne
    ("Р", "к-с"): 5000,         # generic work per-set (small works)
    ("Р", "компл"): 5000,       # generic work per-set
    ("Р", "месяц"): 50000,      # generic monthly overhead
    # Machinery (МХ)
    ("МХ", "маш.-ч"): 3000,     # generic machine-hour
    ("МХ", "маш-ч"): 3000,      # generic machine-hour (alt spelling)
    ("МХ", "м2"): 150,           # generic scaffolding per-m2
}

# Specific material prices for known items without DB match
_SPECIFIC_PRICES: dict[str, float] = {
    # Construction site items (common VOR items without ФССЦ codes)
    "блок-контейнер": 250000,     # ~250K per container unit
    "плиты дорожные": 6000,       # road slabs per piece
    "информационный баннер": 15000,  # info banner
    "ограждение металлическое": 2500,  # per m2
    "штора противопожарная": 80000,  # fire curtain per piece
    "привод электрический": 25000,   # electric drive per piece
    "леса строительные": 150,        # scaffolding per m2
    "кран-манипулятор": 4500,        # crane-manipulator per machine-hour
}

# Administrative/overhead items — these get price=0 with explicit source
_ADMIN_KEYWORDS = [
    "пир", "проектн", "подключен", "подготовит", "содержан", "охран",
    "временн", "изыскательск", "авторский", "технический надзор",
    "технадзор", "согласован", "разрешен", "ввод в эксплуатац",
    "компенсация", "амортизация", "аренда",
    "демобилизац", "демонтаж взис", "внутриплощадочн",
    "прочие расходы", "сигнальное освещен", "временное освещен",
    "бытовой городок", "ограждени", "механизаци",
    "штаб", "мощност",
]


def _build_enhanced_name_index(rp_by_name: dict) -> dict:
    """Build a stemmed name index for better fuzzy matching.

    Returns: {stemmed_first_word -> [entries]} where stemmed form strips
    Russian plural/case endings.
    """
    stemmed = {}
    for first_word, entries in rp_by_name.items():
        stem = _stem_russian(first_word)
        if stem not in stemmed:
            stemmed[stem] = []
        stemmed[stem].extend(entries)
    return stemmed


def _search_by_contains(
    rp_all_list: list,
    search_keywords: list[str],
    comp_unit: str = "",
) -> dict | None:
    """Search all resource prices by keyword containment.

    Returns the best-matching entry or None.
    """
    best = None
    best_score = 0
    for r in rp_all_list:
        rname = (r.get("name") or "").lower()
        score = sum(1 for kw in search_keywords if kw in rname)
        if score > best_score and r["price"] > 0:
            best_score = score
            best = r
    return best


def _validate_composition_quantities(positions: list) -> int:
    """Validate and cap composition item quantities that are unreasonable.

    Rules:
    - Absolute line total cap: 100M per composition line (residential project)
    - Machinery hours: cap at 50,000 per position
    - Material line totals > 50M are capped

    Returns number of capped items.
    """
    from vor.models import CompositionType
    capped = 0

    for pos in positions:
        for pi in pos.items:
            comp = pi.composition
            if pi.unit_price <= 0 or comp.quantity <= 0:
                continue

            line_total = comp.quantity * pi.unit_price

            # Cap A: Absolute line total cap — no single composition line
            # should exceed 100M rubles in a residential project
            if line_total > 100_000_000:
                old_qty = comp.quantity
                comp.quantity = round(100_000_000 / pi.unit_price, 2)
                logger.warning(
                    "Qty capped: %s [%s] %.2f->%.2f (total %.0f->100M)",
                    comp.name[:40], comp.code, old_qty, comp.quantity, line_total,
                )
                capped += 1
                continue

            # Cap B: For machinery, hours should be reasonable
            if comp.type == CompositionType.MACHINERY:
                # More than 50,000 machine-hours is absurd for any single item
                if comp.quantity > 50_000:
                    old_qty = comp.quantity
                    comp.quantity = min(comp.quantity, 10_000)
                    logger.warning(
                        "Mach qty capped: %s %.0f->%.0f",
                        comp.name[:40], old_qty, comp.quantity,
                    )
                    capped += 1

    return capped


async def phase1_extract():
    """Parse VOR, classify, build prompts, save batches."""
    from vor.parser import parse_vor_excel
    from vor.agents.classifier import classify_sections
    from vor.agents.registry import ExpertRegistry
    from vor.agents.expert import COMPOSITION_SYSTEM_PROMPT
    from vor.providers.gesn_sqlite import GesnSqliteProvider

    print(f"=== Phase 1: Extract Batches ===")
    print(f"VOR: {VOR_PATH.name} ({VOR_PATH.stat().st_size // 1024} KB)")

    excel_bytes = VOR_PATH.read_bytes()
    items = parse_vor_excel(excel_bytes)
    print(f"Parsed: {len(items)} items")

    # Classify
    assignments = classify_sections(items)
    print(f"Classified into {len(assignments)} domains:")
    for domain, indices in assignments.items():
        print(f"  {domain.value}: {len(indices)} items")

    # Create batches directory
    BATCHES_DIR.mkdir(exist_ok=True)

    # Create provider for candidate search
    provider = GesnSqliteProvider(GESN_DB)

    # Dummy LLM callback (not used during extraction)
    async def dummy_llm(sys_p, usr_p):
        return ""

    # For each domain, find candidates and build prompts
    registry = ExpertRegistry(
        provider=provider, llm_callback=dummy_llm, project_root=PROJECT_ROOT
    )
    batch_count = 0

    for domain, indices in assignments.items():
        expert = registry.create_expert(domain)
        BATCH_SIZE = 25

        # Find candidates
        candidates = await expert._find_all_candidates(indices, items)
        cand_by_idx = {c["item_idx"]: c for c in candidates}

        # Build batches
        for batch_start in range(0, len(indices), BATCH_SIZE):
            batch_indices = indices[batch_start: batch_start + BATCH_SIZE]

            # Build candidates for this batch (include items without candidates)
            batch_candidates = []
            for idx in batch_indices:
                if idx in cand_by_idx:
                    batch_candidates.append(cand_by_idx[idx])
                else:
                    item = items[idx] if idx < len(items) else None
                    if item:
                        batch_candidates.append({
                            "item_idx": idx,
                            "name": item.name,
                            "unit": item.unit,
                            "section": getattr(item, 'section', ''),
                            "candidates": [],
                        })

            if not batch_candidates:
                continue

            # Build prompts
            system_prompt = expert._build_composition_system_prompt()
            user_prompt = expert._build_composition_user_prompt(batch_candidates, items)

            batch_id = f"{domain.value}_{batch_start:04d}"
            batch_data = {
                "batch_id": batch_id,
                "domain": domain.value,
                "batch_start": batch_start,
                "batch_indices": batch_indices,
                "item_count": len(batch_candidates),
                "system_prompt": system_prompt,
                "user_prompt": user_prompt,
            }

            batch_file = BATCHES_DIR / f"{batch_id}.json"
            batch_file.write_text(json.dumps(batch_data, ensure_ascii=False, indent=2), encoding="utf-8")
            batch_count += 1
            print(f"  Batch {batch_id}: {len(batch_candidates)} items -> {batch_file.name}")

    # Save metadata
    meta = {
        "total_items": len(items),
        "total_batches": batch_count,
        "domains": {d.value: len(idx) for d, idx in assignments.items()},
        "vor_file": str(VOR_PATH),
        "timestamp": time.strftime("%Y-%m-%d %H:%M:%S"),
    }
    (BATCHES_DIR / "_meta.json").write_text(json.dumps(meta, ensure_ascii=False, indent=2), encoding="utf-8")

    print(f"\n=== Phase 1 Complete ===")
    print(f"Total batches: {batch_count}")
    print(f"Saved to: {BATCHES_DIR}")
    print(f"\nNext: Claude processes each batch -> saves to {RESPONSES_DIR}/")


# ═══════════════════════════════════════════════════════════════════════
# Post-pricing corrections: fix systematic price model errors
# ═══════════════════════════════════════════════════════════════════════

# Materials sold in bags/packs where DB stores per-package price, not per-kg
# Format: keyword -> (divisor, unit, max_per_unit_price_after_correction)
_BAG_MATERIALS: dict[str, tuple[int, str, float]] = {
    "клей": (25, "кг", 30),         # 25kg bags, max ~30 руб/кг
    "затирк": (5, "кг", 100),       # 5kg bags, max ~100 руб/кг
    "шпакл": (25, "кг", 40),        # 25kg bags
    "шпатл": (25, "кг", 40),        # alt spelling
    "наливн": (25, "кг", 40),       # self-leveling compound
    "гипс": (25, "кг", 30),         # gypsum mixes
    "штукатурк.*сух": (25, "кг", 40),  # dry plaster mixes
}

# Materials priced as rental, not purchase
# Format: keyword -> (unit, min_price, max_price)
_RENTAL_MATERIALS: dict[str, tuple[str, float, float]] = {
    "опалубк": ("м2", 300, 850),      # formwork rental 300-850 руб/м2/цикл
    "леса строител": ("м2", 100, 200),  # scaffolding rental
}


def _apply_price_corrections(positions: list) -> int:
    """Fix known systematic price errors after main pricing.

    Returns count of corrections applied.
    """
    from vor.models import CompositionType
    from vor.agents.expert import _normalize_unit
    import re as _re

    corrections = 0

    for pos in positions:
        for pi in pos.items:
            comp = pi.composition
            if pi.unit_price <= 0:
                continue
            name_lower = comp.name.lower() if comp.name else ""
            unit_norm = _normalize_unit(comp.unit) if comp.unit else ""

            # Fix 1A: Rental materials (опалубка priced as purchase → rental)
            if comp.type == CompositionType.MATERIAL:
                for keyword, (expected_unit, min_price, max_price) in _RENTAL_MATERIALS.items():
                    if keyword in name_lower and unit_norm == expected_unit:
                        if pi.unit_price > max_price:
                            old = pi.unit_price
                            pi.unit_price = max_price
                            pi.price_source += f" (аренда, было {old:.0f})"
                            corrections += 1
                            logger.info(
                                "Rental correction (high): %s %.0f → %.0f руб/%s",
                                comp.name[:40], old, max_price, comp.unit,
                            )
                        elif pi.unit_price < min_price:
                            old = pi.unit_price
                            pi.unit_price = min_price
                            pi.price_source += f" (аренда мин, было {old:.0f})"
                            corrections += 1
                            logger.info(
                                "Rental correction (low): %s %.0f → %.0f руб/%s",
                                comp.name[:40], old, min_price, comp.unit,
                            )
                        break

            # Fix 1B: Bag materials (клей per-bag → per-kg)
            if comp.type == CompositionType.MATERIAL and unit_norm == "кг":
                for keyword, (divisor, _, max_after) in _BAG_MATERIALS.items():
                    if _re.search(keyword, name_lower):
                        if pi.unit_price > max_after:
                            old = pi.unit_price
                            corrected = round(old / divisor, 2)
                            if corrected <= max_after:
                                pi.unit_price = corrected
                                pi.price_source += f" (/{divisor} мешок)"
                                corrections += 1
                                logger.info(
                                    "Bag correction: %s %.0f / %d = %.2f руб/кг",
                                    comp.name[:40], old, divisor, corrected,
                                )
                            else:
                                # Even after division still too high — use max
                                pi.unit_price = max_after
                                pi.price_source += f" (макс. {max_after})"
                                corrections += 1
                        break

    return corrections


def _supplement_incomplete_compositions(positions: list) -> int:
    """Add missing components to incomplete compositions.

    Fix 1C (фасад НВФ missing insulation/substructure) and
    Fix 1D (плитка missing adhesive/grout).

    Returns count of items added.
    """
    from vor.models import CompositionType, CompositionItem, PricedItem

    supplements = 0

    for pos in positions:
        if not pos.items:
            continue

        # Detect position type from item names and codes
        names = " ".join((pi.composition.name or "").lower() for pi in pos.items)
        pos_unit = ""
        pos_qty = 0.0
        for pi in pos.items:
            if pi.composition.type == CompositionType.WORK:
                pos_unit = pi.composition.unit or ""
                pos_qty = pi.composition.quantity or 0.0
                break

        # Calculate current total per unit
        total_cost = sum(pi.unit_price * (pi.composition.quantity or 0) for pi in pos.items)
        if pos_qty > 0:
            cost_per_unit = total_cost / pos_qty
        else:
            cost_per_unit = total_cost

        # Material names only (for component checks)
        mat_names = " ".join(
            (pi.composition.name or "").lower()
            for pi in pos.items
            if pi.composition.type == CompositionType.MATERIAL
        )

        # Fix 1C: Фасад НВФ — must have подсистема + утеплитель + мембрана + крепёж
        is_nvf = any(kw in names for kw in ["вентилируем", "вентфасад", "нвф", "навесн"])
        if is_nvf and pos_qty > 0:
            # Check what's missing in materials
            has_substructure = any(kw in mat_names for kw in ["подсистем", "каркас"])
            has_insulation = any(kw in mat_names for kw in ["утеплител", "минват", "теплоизоляц"])
            has_membrane = any(kw in mat_names for kw in ["мембран", "ветрозащит"])
            has_fasteners = any(kw in mat_names for kw in ["крепёж", "крепеж", "дюбел", "кляммер"])

            has_cladding = any(kw in mat_names for kw in ["облицов", "керамогранит", "кассет", "фиброцемент", "композит"])

            nvf_supplements = []
            if not has_substructure:
                nvf_supplements.append(("Подсистема НВФ (алюминий)", "м2", 1.0, 1000))
            if not has_insulation:
                nvf_supplements.append(("Утеплитель минвата 100мм", "м2", 1.15, 500))
            if not has_membrane:
                nvf_supplements.append(("Мембрана ветрозащитная", "м2", 1.1, 250))
            if not has_cladding:
                nvf_supplements.append(("Облицовка (керамогранит)", "м2", 1.05, 1200))
            if not has_fasteners:
                nvf_supplements.append(("Крепёж фасадный (дюбели, кляммеры)", "шт", 6.0, 35))

            for name, unit, coeff, price in nvf_supplements:
                qty = pos_qty * coeff
                comp = CompositionItem(
                    type=CompositionType.MATERIAL,
                    code="",
                    name=name,
                    unit=unit,
                    quantity=qty,
                    quantity_formula=f"{pos_qty} × {coeff}",
                )
                pi = PricedItem(
                    composition=comp,
                    unit_price=price,
                    price_source="дополнение НВФ",
                    price_year=2025,
                )
                pos.items.append(pi)
                supplements += 1
                logger.info("NVF supplement: +%s (%.0f %s × %.0f руб)", name, qty, unit, price)

        # Fix 1D: Плиточные работы — must have клей + затирка + грунтовка + плитка
        is_tile = any(kw in names for kw in ["облицовка плит", "плиточн"]) and not is_nvf
        if is_tile and pos_qty > 0:
            has_adhesive = any(kw in mat_names for kw in ["клей"])
            has_grout = any(kw in mat_names for kw in ["затирк"])
            has_primer = any(kw in mat_names for kw in ["грунтовк", "праймер"])
            has_tile = any(kw in mat_names for kw in ["плитк", "керам", "камен", "мозаик"])

            tile_supplements = []
            if not has_tile:
                tile_supplements.append(("Плитка керамическая", "м2", 1.05, 900))
            if not has_adhesive:
                tile_supplements.append(("Клей плиточный", "м2", 1.0, 120))
            if not has_grout:
                tile_supplements.append(("Затирка для швов", "м2", 1.0, 200))
            if not has_primer:
                tile_supplements.append(("Грунтовка", "м2", 1.0, 80))

            for name, unit, coeff, price in tile_supplements:
                qty = pos_qty * coeff
                new_comp = CompositionItem(
                    type=CompositionType.MATERIAL, code="", name=name,
                    unit=unit, quantity=qty, quantity_formula=f"{pos_qty} × {coeff}",
                )
                new_pi = PricedItem(
                    composition=new_comp, unit_price=price,
                    price_source="дополнение плитка", price_year=2025,
                )
                pos.items.append(new_pi)
                supplements += 1

        # Fix: Штукатурные работы — must have штукатурная смесь
        is_plaster = any(kw in names for kw in ["оштукатурив", "штукатурн"]) and not is_nvf and not is_tile
        if is_plaster and pos_qty > 0:
            has_plaster_mix = any(kw in mat_names for kw in ["штукатурк", "смесь", "раствор"])
            if not has_plaster_mix:
                new_comp = CompositionItem(
                    type=CompositionType.MATERIAL, code="", name="Смесь штукатурная",
                    unit="кг", quantity=pos_qty * 8, quantity_formula=f"{pos_qty} × 8 кг/м2",
                )
                new_pi = PricedItem(
                    composition=new_comp, unit_price=12,
                    price_source="дополнение штукатурка", price_year=2025,
                )
                pos.items.append(new_pi)
                supplements += 1

        # Fix: Гидроизоляционные работы — must have гидроизоляционный материал
        is_waterproof = any(kw in names for kw in ["гидроизоляц"]) and not is_nvf
        if is_waterproof and pos_qty > 0:
            has_hydro_mat = any(kw in mat_names for kw in [
                "гидроизоляц", "рулонн", "мастик", "битум", "мембран",
                "наплавл", "пленк", "плёнк", "покрыти", "изоляц", "праймер",
            ])
            if not has_hydro_mat:
                new_comp = CompositionItem(
                    type=CompositionType.MATERIAL, code="",
                    name="Гидроизоляция рулонная наплавляемая",
                    unit="м2", quantity=pos_qty * 1.15, quantity_formula=f"{pos_qty} × 1.15 (нахлёст)",
                )
                new_pi = PricedItem(
                    composition=new_comp, unit_price=400,
                    price_source="дополнение гидроизоляция", price_year=2025,
                )
                pos.items.append(new_pi)
                supplements += 1

    return supplements


def _price_all_in_memory(
    positions: list,
    fer_by_code: dict,
    fer_by_prefix: dict,
    rp_by_code: dict,
    rp_by_name: dict,
    res_units: dict,
    expert,
    rp_by_stem: dict | None = None,
    rp_machinery: list | None = None,
) -> list:
    """Price ALL composition items using in-memory lookups only. ZERO DB calls.

    Args:
        positions: list of PricedPosition to price in-place
        fer_by_code: {code -> {direct_cost, ...}} from fer_prices table
        fer_by_prefix: {prefix5 -> {code, direct_cost, ...}} first match per prefix
        rp_by_code: {code -> {price, measure_unit}} from resource_prices table
        rp_by_name: {first_word -> [{code, name, price, measure_unit}]} for fuzzy name search
        res_units: {code -> measure_unit} from resources table
        expert: ExpertAgent instance (used only for _extract_price_from_encyclopedia)
        rp_by_stem: {stemmed_first_word -> [entries]} enhanced fuzzy index
        rp_machinery: list of all machinery resource_prices entries (code 91.xx)

    Returns:
        Same positions list, mutated with prices filled in.
    """
    from vor.constants import FER_INDEX_2025, FSSC_INDEX_2025, STANDARD_LABOR_RATE
    from vor.agents.expert import _get_unit_multiplier
    from vor.models import CompositionType

    MAX_UNIT_PRICE = 1_000_000  # sanity cap — raised to 1M for base-2000 × 9 indexed prices
    rp_by_stem = rp_by_stem or {}
    rp_machinery = rp_machinery or []

    stats = {"work": 0, "work_priced": 0, "mat": 0, "mat_priced": 0,
             "mach": 0, "mach_priced": 0, "labor": 0, "labor_priced": 0,
             "capped": 0, "stem_match": 0, "alias_match": 0, "contains_match": 0,
             "unit_convert": 0, "llm_price": 0, "typical_price": 0, "specific_price": 0,
             "admin": 0, "code_from_name": 0}

    # Unit cross-conversion factors: (db_unit, comp_unit) -> factor
    # If DB price is per тонна and VOR item is in кг, multiply price by 0.001
    import re as _re
    from vor.agents.expert import _normalize_unit

    def _unit_cross_factor(db_unit: str, comp_unit: str) -> float:
        """Compute conversion factor when DB and composition use different units.

        Returns factor to multiply DB-derived unit price by.
        E.g., DB=т, comp=кг -> 0.001 (divide tonne price by 1000 to get per-kg)
              DB=кг, comp=т -> 1000 (multiply kg price by 1000 to get per-t)
              DB=м3, comp=м3 -> 1.0
        """
        if not db_unit or not comp_unit:
            return 1.0
        db_norm = _normalize_unit(db_unit)
        comp_norm = _normalize_unit(comp_unit)
        if db_norm == comp_norm:
            return 1.0
        # Weight conversions: т <-> кг
        if db_norm == "т" and comp_norm == "кг":
            return 0.001
        if db_norm == "кг" and comp_norm == "т":
            return 1000.0
        # Volume conversions: л <-> м3
        if db_norm == "л" and comp_norm == "м3":
            return 1000.0
        if db_norm == "м3" and comp_norm == "л":
            return 0.001
        # Length: км <-> м
        if db_norm == "км" and comp_norm == "м":
            return 0.001
        if db_norm == "м" and comp_norm == "км":
            return 1000.0
        # Cannot convert — return 1.0 (mismatch stays, capped by MAX_UNIT_PRICE)
        return 1.0

    def _apply_rp_price(rp_entry: dict, comp_code: str, comp_unit: str = "", is_machinery: bool = False) -> tuple[float, str]:
        """Convert a resource_prices entry to a unit price, with unit cross-conversion."""
        raw_unit = rp_entry.get("measure_unit", "")
        if not raw_unit and comp_code and comp_code in res_units:
            raw_unit = res_units[comp_code]
        multiplier = _get_unit_multiplier(raw_unit)
        base_price = round(rp_entry["price"] * FSSC_INDEX_2025 / multiplier, 2)
        # Apply cross-conversion if units differ
        if comp_unit and raw_unit:
            # Strip the quantity prefix from raw_unit for comparison
            # e.g., "100 м2" -> "м2", "1000 л" -> "л"
            stripped_db_unit = _re.sub(r'^\d+\s*', '', raw_unit.strip())
            factor = _unit_cross_factor(stripped_db_unit, comp_unit)
            if factor != 1.0:
                base_price = round(base_price * factor, 2)
                stats["unit_convert"] += 1
        return base_price, raw_unit

    def _score_candidates(candidates: list, terms: list[str], comp_unit: str = "") -> dict | None:
        """Find the best matching candidate by term overlap score.

        Prefers candidates whose measure_unit is compatible with comp_unit.
        Filters out generic prepositions from scoring.
        On ties, prefers the MEDIAN-priced option (avoids both extremes).
        """
        # Remove very common prepositions that inflate scores
        meaningful_terms = [t for t in terms if t not in (
            "для", "из", "на", "при", "без", "под", "над", "тип",
            "или", "все", "всех", "типов", "марка", "марки",
        )]
        if not meaningful_terms:
            meaningful_terms = terms

        comp_unit_norm = _normalize_unit(comp_unit) if comp_unit else ""

        # Collect all scored candidates
        scored: list[tuple[int, float, dict]] = []  # (score, price, entry)

        for r in candidates:
            if r["price"] <= 0:
                continue
            rname = (r.get("name") or "").lower()
            score = sum(1 for t in meaningful_terms if t in rname)
            if score <= 0:
                continue

            # Bonus/penalty for unit compatibility
            r_unit = r.get("measure_unit", "")
            unit_compatible = True
            if comp_unit_norm and r_unit:
                r_unit_stripped = _re.sub(r'^\d+\s*', '', r_unit.strip())
                r_unit_norm = _normalize_unit(r_unit_stripped)
                if r_unit_norm == comp_unit_norm:
                    score += 3  # strong bonus for matching unit
                elif _unit_cross_factor(r_unit_stripped, comp_unit) != 1.0:
                    score += 1  # small bonus if convertible (т<->кг)
                else:
                    # Check for multiplier mismatch (e.g., "10 шт" vs "шт")
                    r_multiplier = _get_unit_multiplier(r_unit)
                    if r_multiplier > 1 and r_unit_norm == comp_unit_norm:
                        score += 2  # same base unit, just different multiplier
                    else:
                        # Truly incompatible units: т vs шт, м3 vs шт, etc.
                        # Mark as incompatible — only use if nothing better
                        unit_compatible = False

            scored.append((score, r["price"], r, unit_compatible))

        if not scored:
            return None

        # Separate unit-compatible and incompatible candidates
        compatible = [(s, p, e) for s, p, e, compat in scored if compat]

        # STRICT: Only use unit-compatible candidates. Incompatible units
        # (e.g., "т" vs "шт") cause catastrophic pricing errors.
        # Better to leave unpriced than to assign a per-tonne price to per-piece.
        pool = compatible
        if not pool:
            return None

        # Find top score
        top_score = max(s for s, _, _ in pool)
        if top_score <= 0:
            return None

        # Among tied top-scorers, pick the MEDIAN price
        # This avoids picking exotic expensive variants (like баритовый песок)
        # while also avoiding unrealistically cheap options
        tied = [(price, entry) for sc, price, entry in pool if sc == top_score]
        tied.sort(key=lambda x: x[0])  # sort by price

        # Pick median
        median_idx = len(tied) // 2
        return tied[median_idx][1]

    def _try_name_match(comp_name: str, comp_code: str, comp_unit: str, is_machinery: bool) -> tuple[float, str]:
        """Enhanced name matching with stemming, aliases, and contains search.

        Returns (unit_price, price_source) or (0.0, "") if not found.
        """
        terms = [w for w in comp_name.lower().split() if len(w) >= 3]
        if not terms:
            return 0.0, ""

        first_term = terms[0]

        # Strategy 1: Exact first-word match (original algorithm)
        candidates = rp_by_name.get(first_term, [])
        if candidates:
            best = _score_candidates(candidates, terms, comp_unit)
            if best:
                price, _ = _apply_rp_price(best, comp_code, comp_unit, is_machinery=is_machinery)
                return price, "ФССЦ (имя)×индекс"

        # Strategy 2: Stemmed first-word match
        stem = _stem_russian(first_term)
        candidates = rp_by_stem.get(stem, [])
        if candidates:
            best = _score_candidates(candidates, terms, comp_unit)
            if best:
                price, _ = _apply_rp_price(best, comp_code, comp_unit, is_machinery=is_machinery)
                stats["stem_match"] += 1
                return price, "ФССЦ (стем)×индекс"

        # Strategy 3: Alias-based search
        alias_map = _MACHINERY_ALIASES if is_machinery else _MATERIAL_ALIASES
        alias_keywords = alias_map.get(first_term, [])
        # Also try with first two words for materials
        if not is_machinery and len(terms) >= 2:
            two_word = f"{terms[0]} {terms[1]}"
            alias_keywords = alias_keywords or _MATERIAL_ALIASES.get(two_word, [])
            # Also try second word alone
            if not alias_keywords:
                alias_keywords = _MATERIAL_ALIASES.get(terms[1], [])

        if alias_keywords:
            # Search the alias keywords against the appropriate pool
            search_pool = rp_machinery if is_machinery else []
            # For materials, collect candidates from all matching stems
            if not is_machinery:
                for ak in alias_keywords:
                    ak_stem = _stem_russian(ak.split()[0]) if ak.split() else ak
                    search_pool.extend(rp_by_stem.get(ak_stem, []))
                    search_pool.extend(rp_by_name.get(ak.split()[0] if ak.split() else ak, []))

            # Score using alias keywords + original terms
            all_search_terms = list(set(terms + alias_keywords))
            best = _score_candidates(search_pool, all_search_terms, comp_unit)
            if best:
                price, _ = _apply_rp_price(best, comp_code, comp_unit, is_machinery=is_machinery)
                stats["alias_match"] += 1
                return price, "ФССЦ (алиас)×индекс"

        # Strategy 4: For machinery — contains-based search across all machinery entries
        if is_machinery and rp_machinery:
            # Search by the main term (stem) in name
            search_terms = [stem] + [_stem_russian(t) for t in terms[1:] if len(t) >= 3]
            best = _score_candidates(rp_machinery, search_terms, comp_unit)
            if best:
                price, _ = _apply_rp_price(best, comp_code, comp_unit, is_machinery=True)
                stats["contains_match"] += 1
                return price, "ФССЦ (маш.поиск)×индекс"

        # Strategy 5: For materials with >=2 terms, try second word as the key
        if not is_machinery and len(terms) >= 2:
            for alt_term in terms[1:3]:
                candidates = rp_by_name.get(alt_term, [])
                if not candidates:
                    candidates = rp_by_stem.get(_stem_russian(alt_term), [])
                if candidates:
                    best = _score_candidates(candidates, terms, comp_unit)
                    if best:
                        price, _ = _apply_rp_price(best, comp_code, comp_unit, is_machinery=False)
                        stats["stem_match"] += 1
                        return price, "ФССЦ (имя2)×индекс"

        # Strategy 6: Try ALL terms as potential index keys (broader search)
        if not is_machinery and len(terms) >= 2:
            # Collect candidates from ALL meaningful terms
            all_candidates = []
            for t in terms:
                if len(t) < 3:
                    continue
                cs = rp_by_name.get(t, [])
                if not cs:
                    cs = rp_by_stem.get(_stem_russian(t), [])
                all_candidates.extend(cs)
            # Also try alias match on all terms
            for t in terms:
                for ak_list in (_MATERIAL_ALIASES.get(t, []),):
                    for ak in ak_list:
                        ak_first = ak.split()[0] if ak.split() else ak
                        all_candidates.extend(rp_by_name.get(ak_first, []))
                        all_candidates.extend(rp_by_stem.get(_stem_russian(ak_first), []))
            if all_candidates:
                best = _score_candidates(all_candidates, terms, comp_unit)
                if best:
                    price, _ = _apply_rp_price(best, comp_code, comp_unit, is_machinery=False)
                    stats["contains_match"] += 1
                    return price, "ФССЦ (широк.поиск)×индекс"

        return 0.0, ""

    # Unit-based sanity cap tables (defined once, used per item)
    _WORK_CAPS = {"м2": 5000, "м3": 20000, "т": 50000, "шт": 100000,
                  "м": 3000, "компл": 500000}
    _MAT_CAPS = {"кг": 2000, "м2": 5000, "м3": 20000, "т": 150000,
                 "шт": 50000, "м": 5000, "л": 1000}
    _MACH_CAPS = {"маш-ч": 10000, "маш.-ч": 10000}

    for pos in positions:
        for pi in pos.items:
            comp = pi.composition
            unit_price = 0.0
            price_source = "не найдена"
            price_year = 0

            if comp.type == CompositionType.WORK:
                stats["work"] += 1
                work_code = comp.code or ""

                # If no code, try to extract from name (e.g. "06-13-003-08 Устройство...")
                if not work_code and comp.name:
                    code_match = _re.match(r'(\d{2}-\d{2}-\d{3}-\d{2})', comp.name.strip())
                    if code_match:
                        work_code = code_match.group(1)

                # Try exact FER code
                if work_code and work_code in fer_by_code:
                    d = fer_by_code[work_code]
                    unit_price = round(d["direct_cost"] * FER_INDEX_2025, 2)
                    price_source = "ФЕР×индекс"
                    price_year = 2025
                elif work_code:
                    # Prefix fallback: try first 5 chars (e.g. "08-02")
                    prefix = work_code[:5]
                    if prefix in fer_by_prefix:
                        d = fer_by_prefix[prefix]
                        unit_price = round(d["direct_cost"] * FER_INDEX_2025, 2)
                        price_source = f"ФЕР×индекс (префикс {prefix})"
                        price_year = 2025

                # Fallback 1: LLM-provided price
                if unit_price == 0.0 and hasattr(pi, '_llm_price') and pi._llm_price > 0:
                    unit_price = pi._llm_price
                    price_source = "LLM"
                    price_year = 2025

                # Fallback 2: Typical price by unit type
                if unit_price == 0.0 and comp.unit:
                    typical = _TYPICAL_PRICES.get(("Р", comp.unit), 0)
                    if typical > 0:
                        unit_price = typical
                        price_source = "типовая (Р)"
                        price_year = 2025

                if unit_price > 0:
                    stats["work_priced"] += 1

            elif comp.type in (CompositionType.MATERIAL, CompositionType.MACHINERY):
                bucket = "mat" if comp.type == CompositionType.MATERIAL else "mach"
                is_machinery = comp.type == CompositionType.MACHINERY
                stats[bucket] += 1

                # Priority 1: Encyclopedia (market prices 2024-2025)
                enc_price, enc_unit = expert._extract_price_from_encyclopedia(comp.name)
                if enc_price is not None and enc_price > 0:
                    # Sanity-check encyclopedia prices against unit-based caps
                    # Encyclopedia can match wrong lines and return inflated prices
                    enc_cap = MAX_UNIT_PRICE
                    unit_norm = _normalize_unit(comp.unit) if comp.unit else ""
                    if unit_norm in ("м2", "м²"):
                        enc_cap = 15_000   # max 15K/m2 (expensive tile/stone)
                    elif unit_norm in ("м3", "м³"):
                        enc_cap = 30_000   # max 30K/m3 (specialty concrete)
                    elif unit_norm == "кг":
                        enc_cap = 2_000    # max 2K/kg
                    elif unit_norm == "м":
                        enc_cap = 10_000   # max 10K/m (cable)
                    elif unit_norm in ("шт", "шт."):
                        enc_cap = 200_000  # max 200K/piece (expensive equipment)
                    elif unit_norm in ("компл", "к-с"):
                        enc_cap = 300_000  # max 300K/set

                    if enc_price <= enc_cap:
                        unit_price = enc_price
                        price_source = "энциклопедия"
                        price_year = 2025
                    else:
                        logger.warning(
                            "Encyclopedia price rejected: %s = %.0f > cap %d (%s)",
                            comp.name[:40], enc_price, enc_cap, comp.unit,
                        )
                        # Fall through to ФССЦ lookup
                        enc_price = None

                if enc_price is None or enc_price <= 0:
                    # Priority 2: resource_prices by code
                    if comp.code and comp.code in rp_by_code:
                        rp = rp_by_code[comp.code]
                        up, _ = _apply_rp_price(rp, comp.code, comp.unit, is_machinery=is_machinery)
                        unit_price = up
                        price_source = "ФССЦ×индекс"
                        price_year = 2025

                    # Priority 3: Enhanced name matching (stemming + aliases + contains)
                    if unit_price == 0.0 and comp.name:
                        up, src = _try_name_match(comp.name, comp.code, comp.unit, is_machinery)
                        if up > 0:
                            unit_price = up
                            price_source = src
                            price_year = 2025

                    # Priority 4: Specific known prices for common items
                    if unit_price == 0.0 and comp.name:
                        name_lower = comp.name.lower()
                        for key, specific_price in _SPECIFIC_PRICES.items():
                            if key in name_lower:
                                unit_price = specific_price
                                price_source = "типовая (специф.)"
                                price_year = 2025
                                break

                    # Priority 5: LLM-provided price fallback
                    if unit_price == 0.0 and hasattr(pi, '_llm_price') and pi._llm_price > 0:
                        unit_price = pi._llm_price
                        price_source = "LLM"
                        price_year = 2025

                    # Priority 6: Typical price by type and unit (last resort)
                    if unit_price == 0.0 and comp.unit:
                        type_marker = "М" if comp.type == CompositionType.MATERIAL else "МХ"
                        typical = _TYPICAL_PRICES.get((type_marker, comp.unit), 0)
                        if typical > 0:
                            unit_price = typical
                            price_source = f"типовая ({type_marker})"
                            price_year = 2025

                if unit_price > 0:
                    stats[bucket + "_priced"] += 1

            elif comp.type == CompositionType.LABOR:
                stats["labor"] += 1
                unit_price = STANDARD_LABOR_RATE
                price_source = "норматив 2025"
                price_year = 2025
                stats["labor_priced"] += 1

            # ── Unit-based sanity caps (smart: try /100 before capping) ──

            if unit_price > 0:
                comp_unit = _normalize_unit(comp.unit) if comp.unit else ""
                if comp.type == CompositionType.WORK:
                    cap = _WORK_CAPS.get(comp_unit, 200000)
                    if unit_price > cap:
                        # Maybe per-100 unit price? Try dividing by 100
                        if unit_price / 100 <= cap:
                            unit_price = round(unit_price / 100, 2)
                            price_source += " (/100)"
                            stats["capped"] += 1
                        else:
                            logger.warning(
                                "Work price capped: %s = %.0f > %d руб/%s",
                                comp.name[:40], unit_price, cap, comp.unit,
                            )
                            unit_price = round(cap * 0.7, 2)
                            price_source += " (capped)"
                            stats["capped"] += 1
                elif comp.type == CompositionType.MATERIAL:
                    cap = _MAT_CAPS.get(comp_unit, 100000)
                    if unit_price > cap:
                        if unit_price / 100 <= cap:
                            unit_price = round(unit_price / 100, 2)
                            price_source += " (/100)"
                            stats["capped"] += 1
                        elif unit_price / 25 <= cap:  # per-bag → per-kg
                            unit_price = round(unit_price / 25, 2)
                            price_source += " (/25)"
                            stats["capped"] += 1
                        else:
                            logger.warning(
                                "Material price capped: %s = %.0f > %d руб/%s",
                                comp.name[:40], unit_price, cap, comp.unit,
                            )
                            unit_price = round(cap * 0.5, 2)
                            price_source += " (capped)"
                            stats["capped"] += 1
                elif comp.type == CompositionType.MACHINERY:
                    cap = _MACH_CAPS.get(comp_unit, 15000)
                    if unit_price > cap:
                        logger.warning(
                            "Machinery price capped: %s = %.0f > %d руб/%s",
                            comp.name[:40], unit_price, cap, comp.unit,
                        )
                        unit_price = round(cap * 0.7, 2)
                        price_source += " (capped)"
                        stats["capped"] += 1

            pi.unit_price = unit_price
            pi.price_source = price_source
            pi.price_year = price_year

    # ── Post-pricing corrections: fix known price model errors ──────────
    corrections = _apply_price_corrections(positions)
    stats["price_corrections"] = corrections

    # ── Composition supplements: add missing components ──────────────────
    supplements = _supplement_incomplete_compositions(positions)
    stats["supplements"] = supplements

    # Validate composition quantities (cap absurd totals)
    qty_capped = _validate_composition_quantities(positions)
    stats["qty_capped"] = qty_capped

    # Log summary
    logger.info(
        "Pricing stats: W=%d/%d M=%d/%d MH=%d/%d L=%d/%d capped=%d "
        "stem=%d alias=%d contains=%d unit_conv=%d qty_capped=%d "
        "llm=%d typical=%d specific=%d admin=%d code_from_name=%d "
        "corrections=%d supplements=%d",
        stats["work_priced"], stats["work"],
        stats["mat_priced"], stats["mat"],
        stats["mach_priced"], stats["mach"],
        stats["labor_priced"], stats["labor"],
        stats["capped"],
        stats["stem_match"], stats["alias_match"], stats["contains_match"],
        stats["unit_convert"], qty_capped,
        stats["llm_price"], stats["typical_price"], stats["specific_price"],
        stats["admin"], stats["code_from_name"],
        corrections, supplements,
    )

    return positions


async def phase3_assemble():
    """Read Claude responses, price compositions, assemble Excel.

    Fully in-memory pricing: preloads ALL price data from DB once,
    then prices every item without any further DB calls.
    """
    from vor.parser import parse_vor_excel
    from vor.agents.classifier import classify_sections
    from vor.agents.registry import ExpertRegistry
    from vor.agents.expert import ExpertAgent, _safe_float
    from vor.assembler import VorAssembler
    from vor.models import (
        ExpertDomain, PricedSection, PricedPosition, PricedItem,
        CompositionItem, CompositionType, VerificationReport,
    )

    t0 = time.time()
    print(f"=== Phase 3: Assemble from Responses ===")

    if not RESPONSES_DIR.exists():
        print(f"ERROR: {RESPONSES_DIR} does not exist. Run Phase 2 first.")
        return

    excel_bytes = VOR_PATH.read_bytes()
    items = parse_vor_excel(excel_bytes)
    assignments = classify_sections(items)

    # ── Preload ALL price data into memory (one-time DB access) ──────────
    import sqlite3
    print("  Preloading all prices into memory...")
    t_load = time.time()
    conn = sqlite3.connect(GESN_DB)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    # 1. Load resource_prices (material/machinery pricing)
    cur.execute("SELECT code, name, price, COALESCE(measure_unit, '') as mu FROM resource_prices WHERE price > 0")
    _rp_by_name = {}  # first_word -> [entries]
    _rp_by_code = {}  # code -> {price, measure_unit}
    _rp_machinery = []  # all machinery entries (code 91.xx) for contains-search
    rp_count = 0
    for row in cur.fetchall():
        code, name, price, mu = row["code"], row["name"], row["price"], row["mu"]
        entry = {"code": code, "name": name, "price": price, "measure_unit": mu}
        _rp_by_code[code] = {"price": price, "measure_unit": mu}
        if name:
            first_word = name.lower().split()[0] if name.split() else ""
            if first_word and len(first_word) >= 3:
                if first_word not in _rp_by_name:
                    _rp_by_name[first_word] = []
                _rp_by_name[first_word].append(entry)
        # Collect machinery entries for contains-based search
        if code.startswith("91."):
            _rp_machinery.append(entry)
        rp_count += 1

    # Build enhanced stemmed name index
    _rp_by_stem = _build_enhanced_name_index(_rp_by_name)
    print(f"  Name index: {len(_rp_by_name)} buckets, stemmed: {len(_rp_by_stem)} buckets, "
          f"machinery pool: {len(_rp_machinery)} entries")

    # 2. Load FER prices (work pricing)
    cur.execute("SELECT code, direct_cost FROM fer_prices WHERE direct_cost > 0")
    _fer_by_code = {}
    _fer_by_prefix = {}
    for row in cur.fetchall():
        code, dc = row["code"], row["direct_cost"]
        _fer_by_code[code] = {"direct_cost": dc}
        prefix = code[:5]
        if prefix not in _fer_by_prefix:
            _fer_by_prefix[prefix] = {"direct_cost": dc}

    # 3. Load resource units from resources table (for unit multiplier)
    cur.execute("SELECT code, measure_unit FROM resources WHERE measure_unit IS NOT NULL AND measure_unit != ''")
    _res_units = {}
    for row in cur.fetchall():
        _res_units[row["code"]] = row["measure_unit"]

    conn.close()
    print(f"  Loaded {rp_count} resource prices, {len(_fer_by_code)} FER prices, "
          f"{len(_res_units)} resource units in {time.time()-t_load:.1f}s")

    # ── Create expert agents (for encyclopedia + parse only, NO DB pricing) ─
    async def dummy_llm(sys_p, usr_p):
        return ""

    from vor.providers.gesn_sqlite import GesnSqliteProvider
    provider = GesnSqliteProvider(GESN_DB)
    registry = ExpertRegistry(
        provider=provider, llm_callback=dummy_llm, project_root=PROJECT_ROOT
    )

    # ── Process each domain ─────────────────────────────────────────────
    all_sections: list[PricedSection] = []
    total_positions = 0

    for domain, indices in assignments.items():
        expert = registry.create_expert(domain)
        domain_positions: list[PricedPosition] = []

        # Find all response files for this domain
        response_files = sorted(RESPONSES_DIR.glob(f"{domain.value}_*.json"))
        if not response_files:
            print(f"  {domain.value}: no response files found, skipping")
            all_sections.append(PricedSection(
                domain=domain,
                positions=[],
                verification=VerificationReport(
                    red_flags=[f"Нет ответов для домена {domain.value}"],
                    passed=False,
                ),
            ))
            continue

        for resp_file in response_files:
            resp_data = json.loads(resp_file.read_text(encoding="utf-8"))
            batch_id = resp_data.get("batch_id", resp_file.stem)
            batch_indices = resp_data.get("batch_indices", [])
            llm_response = resp_data.get("response", "")

            if not llm_response:
                print(f"  {batch_id}: empty response, skipping")
                continue

            # Parse the composition response
            parsed = expert._parse_composition_response(llm_response)
            positions_data = parsed.get("positions", [])

            # Extract positions
            for pos_data in positions_data:
                item_idx = pos_data.get("item_idx")
                if item_idx is None:
                    if len(batch_indices) == 1:
                        item_idx = 0
                    else:
                        continue

                if isinstance(item_idx, str):
                    try:
                        item_idx = int(item_idx)
                    except ValueError:
                        continue

                if item_idx < len(batch_indices):
                    abs_idx = batch_indices[item_idx]
                else:
                    continue

                if abs_idx < 0 or abs_idx >= len(items):
                    continue

                row_num = items[abs_idx].row_num
                if not isinstance(row_num, int) or row_num <= 0:
                    continue

                # Build composition items
                composition_items = pos_data.get("composition", [])
                priced_items = []
                for comp_data in composition_items:
                    comp_type_str = comp_data.get("type", "work").lower()
                    try:
                        comp_type = CompositionType(comp_type_str)
                    except ValueError:
                        comp_type = CompositionType.WORK

                    comp_item = CompositionItem(
                        type=comp_type,
                        code=comp_data.get("code", ""),
                        name=comp_data.get("name", ""),
                        unit=comp_data.get("unit", ""),
                        quantity=_safe_float(comp_data.get("quantity", 0)),
                        quantity_formula=comp_data.get("quantity_formula", ""),
                    )
                    pi = PricedItem(
                        composition=comp_item,
                        unit_price=0.0,
                        price_source="",
                        price_year=0,
                    )
                    # Store LLM-provided unit_price hint (if present in response)
                    llm_price = comp_data.get("unit_price", 0)
                    pi._llm_price = float(llm_price) if llm_price else 0.0
                    priced_items.append(pi)

                confidence_str = str(pos_data.get("confidence", "MEDIUM")).upper()
                conf_val = {"HIGH": 0.85, "LOW": 0.30}.get(confidence_str, 0.55)

                domain_positions.append(PricedPosition(
                    original_idx=row_num,
                    items=priced_items,
                    confidence=conf_val,
                    notes=pos_data.get("notes", ""),
                ))

            print(f"  {batch_id}: {len(positions_data)} positions parsed")

        # ── Price all positions in-memory (ZERO DB calls) ──────────────
        if domain_positions:
            _price_all_in_memory(
                domain_positions,
                _fer_by_code, _fer_by_prefix,
                _rp_by_code, _rp_by_name, _res_units,
                expert,
                rp_by_stem=_rp_by_stem,
                rp_machinery=_rp_machinery,
            )
            verification = expert._verify_result(domain_positions)
        else:
            verification = VerificationReport(
                red_flags=[f"Пусто для {domain.value}"],
                passed=False,
            )

        total_positions += len(domain_positions)
        all_sections.append(PricedSection(
            domain=domain,
            positions=domain_positions,
            verification=verification,
        ))
        print(f"  {domain.value}: {len(domain_positions)} priced, "
              f"coverage={verification.coverage_pct:.0f}%, passed={verification.passed}")

    print(f"\nTotal positions priced: {total_positions}")

    # ── Validation ──────────────────────────────────────────────────────
    from vor.validator import VorValidator
    validator = VorValidator()
    validation = validator.validate_all(all_sections)
    report_text = validation.summary()
    print(f"\n{'='*60}")
    print(report_text)
    print(f"{'='*60}")
    # Save report to file (avoids console encoding issues)
    report_path = Path(PROJECT_ROOT) / "vor_validation_report.txt"
    report_path.write_text(report_text, encoding="utf-8")
    print(f"Validation report saved to: {report_path}")

    # ── Assemble Excel ──────────────────────────────────────────────────
    if total_positions > 0:
        assembler = VorAssembler()
        result_bytes = assembler.assemble(excel_bytes, all_sections)
        OUTPUT_PATH.write_bytes(result_bytes)
        print(f"\nOutput: {OUTPUT_PATH} ({len(result_bytes):,} bytes)")

        # Quick stats
        import openpyxl
        wb = openpyxl.load_workbook(OUTPUT_PATH)
        ws = wb.active
        formula_count = sum(
            1 for row in ws.iter_rows(min_row=2, max_row=ws.max_row)
            for cell in row
            if isinstance(cell.value, str) and cell.value.startswith("=")
        )
        print(f"Rows: {ws.max_row}, Formulas: {formula_count}")
        wb.close()

        elapsed = time.time() - t0
        print(f"\nTotal time: {elapsed:.1f}s")
        print(f"\n+++ SUCCESS +++")
    else:
        print("\n!!! No positions priced. Check response files. !!!")


if __name__ == "__main__":
    cmd = sys.argv[1] if len(sys.argv) > 1 else "extract"
    if cmd == "extract":
        asyncio.run(phase1_extract())
    elif cmd == "assemble":
        asyncio.run(phase3_assemble())
    else:
        print(f"Usage: python vor_claude_runner.py [extract|assemble]")
