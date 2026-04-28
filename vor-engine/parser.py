"""Parse an uploaded Excel ВОР (Bill of Quantities) file.

Handles common Russian ВОР formats: detects headers, sections, items,
and extracts name, unit, quantity for each line item.
"""

from __future__ import annotations

import io
import logging
import re
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

from vor.models import VorItem

logger = logging.getLogger(__name__)

# Patterns used for header detection (lowercase).
_HEADER_PATTERNS: dict[str, list[str]] = {
    "num": ["№", "n", "п/п", "№ п/п", "номер", "no"],
    "name": [
        "наименование",
        "название",
        "описание",
        "наименование работ",
        "наименование работ и затрат",
        "виды работ",
        "работы",
    ],
    "unit": ["ед.", "ед. изм.", "ед.изм.", "единица", "единица измерения", "ед"],
    "quantity": ["кол-во", "кол.", "количество", "объем", "объём", "кол"],
    "total": ["итоговая сумма", "итоговая стоимость", "итоговая"],
}

# Minimum columns that must be detected for a valid header row.
_REQUIRED_COLUMNS = {"name", "unit"}


def parse_vor_excel(file_bytes: bytes) -> list[VorItem]:
    """Parse a ВОР Excel file and return a list of VorItem.

    Args:
        file_bytes: Raw bytes of the .xlsx file.

    Returns:
        List of parsed VOR items. Empty list if file is invalid or empty.
    """
    try:
        wb = load_workbook(
            filename=io.BytesIO(file_bytes),
            read_only=True,
            data_only=True,
        )
    except Exception:
        logger.warning("Failed to open workbook — returning empty list")
        return []

    try:
        ws = wb.active
        if ws is None:
            return []

        header_row, col_map = _find_header_row(ws)
        if header_row < 0:
            logger.warning("No header row detected — returning empty list")
            return []

        return _extract_items(ws, header_row, col_map)
    except Exception:
        logger.exception("Unexpected error while parsing VOR Excel")
        return []
    finally:
        try:
            wb.close()
        except Exception:
            pass


def _find_header_row(ws) -> tuple[int, dict[str, int]]:
    """Scan first 30 rows to find the header row.

    Returns:
        (row_number, column_mapping) where column_mapping maps
        logical names ("name", "unit", "quantity", "num") to
        zero-based column indices.  Returns (-1, {}) if not found.
    """
    max_scan = 30
    for row_idx, row in enumerate(ws.iter_rows(max_row=max_scan, values_only=False), start=1):
        col_map = _match_header_row(row)
        if _REQUIRED_COLUMNS.issubset(col_map.keys()):
            logger.debug("Header detected at row %d: %s", row_idx, col_map)
            return row_idx, col_map
    return -1, {}


def _match_header_row(row: tuple) -> dict[str, int]:
    """Try to match cells in a row to known header patterns.

    Returns a mapping of logical column name -> zero-based column index.
    """
    col_map: dict[str, int] = {}
    for col_idx, cell in enumerate(row):
        text = _cell_value_str(cell).strip()
        if not text:
            continue
        text_lower = text.lower()
        for logical_name, patterns in _HEADER_PATTERNS.items():
            if logical_name in col_map:
                continue  # already matched
            for pat in patterns:
                if pat == text_lower or text_lower.startswith(pat):
                    col_map[logical_name] = col_idx
                    break
    return col_map


def _extract_items(ws, header_row: int, col_map: dict[str, int]) -> list[VorItem]:
    """Walk rows below the header and extract VOR items."""
    items: list[VorItem] = []
    current_section = ""
    name_col = col_map["name"]
    unit_col = col_map["unit"]
    qty_col = col_map.get("quantity")
    num_col = col_map.get("num")

    for row_idx, row in enumerate(
        ws.iter_rows(min_row=header_row + 1, values_only=False),
        start=header_row + 1,
    ):
        # Build raw_data dict for reference.
        raw = _row_to_dict(row, col_map)

        name_text = _clean_text(_safe_cell(row, name_col))
        unit_text = _clean_text(_safe_cell(row, unit_col))

        # Skip completely empty rows.
        if not name_text and not unit_text:
            continue

        # Section header detection.
        if _is_section_header(row, col_map, name_text, unit_text):
            current_section = name_text
            continue

        # Skip rows without a name.
        if not name_text:
            continue

        # Skip subtotal / total rows (numeric-only name or keywords).
        if _is_subtotal_row(name_text):
            continue

        quantity = _parse_quantity(_safe_cell(row, qty_col) if qty_col is not None else None)

        items.append(
            VorItem(
                row_num=row_idx,
                name=name_text,
                unit=unit_text if unit_text else "",
                quantity=quantity,
                section=current_section,
                raw_data=raw,
            )
        )

    return items


def _is_section_header(
    row: tuple,
    col_map: dict[str, int],
    name_text: str,
    unit_text: str,
) -> bool:
    """Detect section header rows.

    Section headers typically have a name but no unit of measurement and
    no quantity.  They may also be bold or span merged cells.
    """
    if not name_text:
        return False
    # If there is a unit, it is not a section header.
    if unit_text:
        return False
    # If the "num" column contains something that looks like a section
    # number (e.g., "1", "I", "Раздел 2") — treat it as a section.
    num_col = col_map.get("num")
    qty_col = col_map.get("quantity")
    # No quantity present.
    qty_val = _safe_cell(row, qty_col) if qty_col is not None else None
    if qty_val is not None and _parse_quantity(qty_val) is not None:
        return False
    return True


def _is_subtotal_row(name_text: str) -> bool:
    """Return True if the row looks like a subtotal / total."""
    lower = name_text.lower().strip()
    if re.match(r"^[\d\s.,]+$", lower):
        return True
    subtotal_keywords = ("итого", "всего", "итого по разделу", "total", "subtotal")
    for kw in subtotal_keywords:
        if lower.startswith(kw):
            return True
    return False


def _parse_quantity(value: Any) -> float | None:
    """Parse a cell value into a float quantity, or None."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace("\u00a0", "").replace(" ", "")
    # Handle European format: 1.500,5 or 1500,5
    if "," in text:
        # If both . and , present: . is thousands sep, , is decimal
        if "." in text:
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", ".")
    if not text:
        return None
    try:
        return float(text)
    except (ValueError, TypeError):
        return None


def _clean_text(value: Any) -> str:
    """Normalize a cell value to a clean string."""
    if value is None:
        return ""
    text = str(value).strip()
    # Collapse multiple whitespace into single space.
    text = re.sub(r"\s+", " ", text)
    return text


def _cell_value_str(cell: Any) -> str:
    """Get cell value as a string, handling merged cells and None."""
    if cell is None:
        return ""
    if isinstance(cell, MergedCell):
        return ""
    # ReadOnlyCell (from read_only=True) is not a subclass of Cell,
    # so we use hasattr instead of isinstance.
    val = cell.value if hasattr(cell, "value") else cell
    if val is None:
        return ""
    return str(val)


def _safe_cell(row: tuple, col_idx: int | None) -> Any:
    """Safely get a cell value by column index."""
    if col_idx is None or col_idx >= len(row):
        return None
    cell = row[col_idx]
    if isinstance(cell, MergedCell):
        return None
    # ReadOnlyCell has .value but is not a subclass of Cell.
    if hasattr(cell, "value"):
        return cell.value
    return cell


def _row_to_dict(row: tuple, col_map: dict[str, int]) -> dict:
    """Convert a row to a dict keyed by logical column names."""
    result: dict[str, Any] = {}
    for name, idx in col_map.items():
        result[name] = _safe_cell(row, idx)
    return result
