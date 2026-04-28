"""Assemble a priced VOR Excel from the original workbook + expert results.

Opens the **original** Excel (preserving all rows, formatting, merged cells),
then inserts composition rows below each priced position and fills price /
formula / comment columns.  The result is returned as raw bytes.

Column mapping (from the test VOR "ЖК Сокольники"):

    Col A (1):  "Номер позиции"
    Col B (2):  "№ п/п"
    Col F (6):  "Наименование"
    Col G (7):  "Ед. изм."
    Col H (8):  "Количество заказчика"
    Col O (15): "Цена за единицу"        — WE FILL
    Col P (16): "Итоговая сумма"          — WE FILL (formula)
    Col S (19): "Примечание ГП"           — WE FILL
"""

from __future__ import annotations

import io
import logging
from copy import copy
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from vor.models import (
    CompositionType,
    PricedItem,
    PricedPosition,
    PricedSection,
)

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Column indices (1-based) — loaded from vor_config.yaml if available
# ---------------------------------------------------------------------------

from pathlib import Path as _Path

import yaml as _yaml

def _load_column_config() -> dict:
    """Load column mapping from vor_config.yaml, fallback to defaults."""
    cfg_path = _Path(__file__).parent / "vor_config.yaml"
    defaults = {
        "position_num": 1, "ordinal": 2, "name": 6, "unit": 7,
        "quantity": 8, "unit_price": 15, "total": 16, "comment": 19,
    }
    if cfg_path.exists():
        try:
            with open(cfg_path, "r", encoding="utf-8") as f:
                cfg = _yaml.safe_load(f) or {}
            columns = cfg.get("columns", {})
            for k, v in columns.items():
                if k in defaults:
                    defaults[k] = v
        except Exception:
            pass
    return defaults

_cols = _load_column_config()

COL_POSITION_NUM = _cols["position_num"]  # A — hierarchical position number
COL_ORDINAL = _cols["ordinal"]            # B — ordinal / type marker
COL_NAME = _cols["name"]                  # F — description
COL_UNIT = _cols["unit"]                  # G — unit of measurement
COL_QUANTITY = _cols["quantity"]           # H — quantity
COL_UNIT_PRICE = _cols["unit_price"]      # O — unit price
COL_TOTAL = _cols["total"]                # P — total (formula)
COL_COMMENT = _cols["comment"]            # S — our comment

# Dynamic column letters for formulas
_QTY_LETTER = get_column_letter(COL_QUANTITY)      # "H"
_PRICE_LETTER = get_column_letter(COL_UNIT_PRICE)  # "O"
_TOTAL_LETTER = get_column_letter(COL_TOTAL)        # "P"

# ---------------------------------------------------------------------------
# Type markers written into Col B for inserted rows
# ---------------------------------------------------------------------------

_TYPE_MARKERS: dict[CompositionType, str] = {
    CompositionType.WORK: "Р",
    CompositionType.MATERIAL: "М",
    CompositionType.MACHINERY: "МХ",
    CompositionType.LABOR: "ТР",
}

# ---------------------------------------------------------------------------
# Fills per composition type
# ---------------------------------------------------------------------------

_FILL_WORK = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
_FILL_MATERIAL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
_FILL_MACHINERY = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
_FILL_LABOR = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
_FILL_TOTAL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

_TYPE_FILLS: dict[CompositionType, PatternFill] = {
    CompositionType.WORK: _FILL_WORK,
    CompositionType.MATERIAL: _FILL_MATERIAL,
    CompositionType.MACHINERY: _FILL_MACHINERY,
    CompositionType.LABOR: _FILL_LABOR,
}

_FONT_BOLD = Font(bold=True)
_FONT_NORMAL = Font(bold=False)
_FONT_TOTAL = Font(bold=True)

# FIX 9: Only format columns that contain data (performance optimization)
_DATA_COLS = [COL_ORDINAL, COL_NAME, COL_UNIT, COL_QUANTITY, COL_UNIT_PRICE, COL_TOTAL, COL_COMMENT]


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------


class VorAssembler:
    """Assembles a priced VOR from the original Excel + expert results."""

    def assemble(
        self,
        original_bytes: bytes,
        expert_results: list[PricedSection],
        sheet_name: str | None = None,
    ) -> bytes:
        """Return a new Excel workbook (as bytes) with prices filled in.

        Parameters
        ----------
        original_bytes:
            Raw bytes of the original .xlsx file.
        expert_results:
            List of ``PricedSection`` objects produced by expert agents.
        sheet_name:
            Name of the worksheet to modify.  ``None`` → first sheet.

        Returns
        -------
        bytes
            The modified .xlsx workbook.
        """
        if not expert_results:
            # Nothing to do — return the original unchanged.
            return original_bytes

        wb = load_workbook(filename=io.BytesIO(original_bytes))
        ws: Worksheet = wb[sheet_name] if sheet_name else wb.active  # type: ignore[assignment]

        col_warnings = self._validate_columns(ws)
        for w in col_warnings:
            logger.warning("Column validation: %s", w)

        # 1. Build a flat map: original_row_num → PricedPosition
        row_map = self._build_row_map(expert_results)
        if not row_map:
            return original_bytes

        # 2. Sort positions by original row number DESCENDING (bottom-up insertion)
        sorted_rows = sorted(row_map.keys(), reverse=True)

        # 3. TWO-PASS ALGORITHM:
        #    Pass 1: Insert all rows bottom-up, fill data (no formulas).
        #    Pass 2: Write all formulas using FINAL row positions.
        #
        #    This avoids the bug where formulas reference pre-shift row numbers.

        # Track where each position ended up after all shifts.
        # Each insertion below a row doesn't affect rows above it (bottom-up).
        # But we need the CUMULATIVE shift for formula pass.
        #
        # Strategy: track total rows inserted ABOVE each original row.
        # After all insertions, final_row = orig_row + total_inserted_above.

        # Pass 1: Insert rows and fill data (NOT formulas)
        # Since we go bottom-up, we track how many rows were inserted in total.
        # positions_info will store final locations for pass 2.
        positions_info: list[dict] = []  # [{orig_row, position, final_row, inserts}]
        total_inserted = 0

        for orig_row in sorted_rows:
            position = row_map[orig_row]
            items = position.items
            n_insert = len(items) if items else 0

            if n_insert > 0:
                self._safe_insert_rows(ws, orig_row + 1, n_insert)

                # Fill data on inserted rows (quantities, names, prices — NOT formulas)
                for i, priced_item in enumerate(items):
                    insert_row = orig_row + 1 + i
                    self._fill_inserted_row_data(ws, insert_row, priced_item)

            total_inserted += n_insert
            positions_info.append({
                "orig_row": orig_row,
                "position": position,
                "n_insert": n_insert,
            })

        # Pass 2: Write all formulas using FINAL row positions.
        # After all bottom-up insertions, rows shifted down by the count of
        # rows inserted ABOVE them. Walk positions in ASCENDING order and
        # accumulate the shift.
        positions_info.reverse()  # now ascending by orig_row
        cumulative_shift = 0
        for info in positions_info:
            orig_row = info["orig_row"]
            position = info["position"]
            n_insert = info["n_insert"]
            final_row = orig_row + cumulative_shift

            if n_insert == 0:
                self._write_formula_no_items(ws, final_row, position)
            else:
                first_insert = final_row + 1
                last_insert = final_row + n_insert
                # Formulas on inserted rows
                for i in range(n_insert):
                    r = first_insert + i
                    ws.cell(row=r, column=COL_TOTAL,
                            value=f"={_QTY_LETTER}{r}*{_PRICE_LETTER}{r}")
                # SUM formula on original row
                ws.cell(row=final_row, column=COL_TOTAL,
                        value=f"=SUM({_TOTAL_LETTER}{first_insert}:{_TOTAL_LETTER}{last_insert})")
                # Comment on original row
                comment = _build_comment(position)
                ws.cell(row=final_row, column=COL_COMMENT, value=comment)
                # Bold original row
                for col in (COL_NAME, COL_UNIT, COL_QUANTITY, COL_UNIT_PRICE, COL_TOTAL, COL_COMMENT):
                    ws.cell(row=final_row, column=col).font = _FONT_BOLD

            cumulative_shift += n_insert

        # 5. Save and return bytes
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.read()

    # ------------------------------------------------------------------
    # Internal helpers
    # ------------------------------------------------------------------

    @staticmethod
    def _safe_insert_rows(ws: Worksheet, row: int, amount: int) -> None:
        """Insert rows while preserving merged cell integrity.

        Handles two categories of merged cells:
        1. merges_below: start at or below insertion point → shift both ends down
        2. merges_spanning: start above, end at/below insertion → extend end down
        """
        merges_below = []
        merges_spanning = []
        for merge_range in list(ws.merged_cells.ranges):
            if merge_range.min_row >= row:
                merges_below.append(merge_range)
            elif merge_range.max_row >= row:
                # Merge spans the insertion point
                merges_spanning.append(merge_range)

        # Unmerge both categories before inserting rows
        for mr in merges_below:
            ws.unmerge_cells(str(mr))
        for mr in merges_spanning:
            ws.unmerge_cells(str(mr))

        ws.insert_rows(row, amount=amount)

        # Re-merge below: shift both min and max down
        for mr in merges_below:
            ws.merge_cells(
                start_row=mr.min_row + amount,
                start_column=mr.min_col,
                end_row=mr.max_row + amount,
                end_column=mr.max_col,
            )
        # Re-merge spanning: keep min_row, extend max_row
        for mr in merges_spanning:
            ws.merge_cells(
                start_row=mr.min_row,
                start_column=mr.min_col,
                end_row=mr.max_row + amount,
                end_column=mr.max_col,
            )

    @staticmethod
    def _validate_columns(ws: Worksheet, header_row: int = 1) -> list[str]:
        """Validate that expected columns are in expected positions.

        Returns list of warning strings. Does not fail — just informational.
        """
        warnings: list[str] = []
        expected = {
            COL_NAME: "наименован",
            COL_UNIT: "ед",
            COL_QUANTITY: "кол",
        }
        for col, expected_substr in expected.items():
            cell_val = str(ws.cell(row=header_row, column=col).value or "").lower()
            if expected_substr not in cell_val:
                warnings.append(
                    f"Column {col}: expected '{expected_substr}' in header, "
                    f"got '{ws.cell(row=header_row, column=col).value}'"
                )
        return warnings

    @staticmethod
    def _build_row_map(
        expert_results: list[PricedSection],
    ) -> dict[int, PricedPosition]:
        """Build {excel_row_number: PricedPosition} from all sections."""
        row_map: dict[int, PricedPosition] = {}
        for section in expert_results:
            for position in section.positions:
                idx = position.original_idx
                if idx > 0:
                    row_map[idx] = position
        return row_map

    # _fill_original_row and _fill_original_row_no_items removed.
    # Their logic is now in the two-pass algorithm in assemble() (Pass 2).

    @staticmethod
    def _fill_inserted_row_data(
        ws: Worksheet,
        row: int,
        priced_item: PricedItem,
    ) -> None:
        """Fill data on an inserted row (Pass 1 — NO formulas, those come in Pass 2)."""
        comp = priced_item.composition
        comp_type = comp.type
        marker = _TYPE_MARKERS.get(comp_type, "?")
        fill = _TYPE_FILLS.get(comp_type, _FILL_WORK)

        # Col B: type marker
        ws.cell(row=row, column=COL_ORDINAL, value=marker)

        # Col F: code + name
        name = f"{comp.code} {comp.name}".strip() if comp.code else comp.name
        ws.cell(row=row, column=COL_NAME, value=name)

        # Col G: unit
        ws.cell(row=row, column=COL_UNIT, value=comp.unit)

        # Col H: quantity
        ws.cell(row=row, column=COL_QUANTITY, value=comp.quantity)

        # Col O: unit price
        ws.cell(row=row, column=COL_UNIT_PRICE, value=priced_item.unit_price)

        # Formulas (Col P) are written in Pass 2 after all row shifts are done.

        # Apply fill only to columns with data
        for col in _DATA_COLS:
            cell = ws.cell(row=row, column=col)
            cell.fill = fill
            cell.font = _FONT_NORMAL

    @staticmethod
    def _write_formula_no_items(
        ws: Worksheet,
        row: int,
        position: PricedPosition,
    ) -> None:
        """Write formula for original row with no composition sub-items (Pass 2)."""
        for col in (COL_NAME, COL_UNIT, COL_QUANTITY, COL_UNIT_PRICE, COL_TOTAL, COL_COMMENT):
            ws.cell(row=row, column=col).font = _FONT_BOLD
        ws.cell(row=row, column=COL_TOTAL,
                value=f"={_QTY_LETTER}{row}*{_PRICE_LETTER}{row}")
        comment = _build_comment(position)
        ws.cell(row=row, column=COL_COMMENT, value=comment)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _build_comment(position: PricedPosition) -> str:
    """Build a comment string for Col S from position metadata."""
    parts: list[str] = []

    if position.notes:
        parts.append(position.notes)

    # Collect GESN codes from work items
    gesn_codes = []
    for item in position.items:
        if item.composition.type == CompositionType.WORK and item.composition.code:
            gesn_codes.append(item.composition.code)
    if gesn_codes:
        parts.append(f"ГЭСН: {', '.join(gesn_codes)}")

    # Confidence
    if position.confidence > 0:
        parts.append(f"Уверенность: {position.confidence:.0%}")

    # Price sources
    sources = {item.price_source for item in position.items if item.price_source}
    if sources:
        parts.append(f"Источник: {', '.join(sorted(sources))}")

    # Iterative validation metadata
    if position.expert_comment:
        parts.append(f"[Эксперт] {position.expert_comment}")
    if position.validation_errors:
        parts.append(f"[Ошибки] {'; '.join(position.validation_errors[:3])}")
    if position.iteration > 0:
        parts.append(f"[Итерация {position.iteration}]")

    return "; ".join(parts) if parts else ""
