# -*- coding: utf-8 -*-
"""
Автозаполнение пустого ВОР шаблонами (v3).
Синхронизирован с JS-логикой (vorMatcher.js + vorExcelGenerator.js).

Usage:
    python fill_vor.py input.xlsx [output.xlsx]
"""
import sys
import re
import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from pathlib import Path

sys.stdout.reconfigure(encoding='utf-8')

FILL_HEADER   = PatternFill('solid', fgColor='E0E0E0')
FILL_SECTION  = PatternFill('solid', fgColor='FFF2CC')
FILL_POSITION = PatternFill('solid', fgColor='FFCCCC')
FILL_WORK     = PatternFill('solid', fgColor='E6D9F2')
FILL_MATERIAL = PatternFill('solid', fgColor='E8F5E0')
THIN_BORDER = Border(
    left=Side('thin', 'CCCCCC'), right=Side('thin', 'CCCCCC'),
    top=Side('thin', 'CCCCCC'), bottom=Side('thin', 'CCCCCC'))

# ─── Шаблоны ─────────────────────────────────────────────────────────
TEMPLATES = {
    'spk_profile': {
        'costPath': 'ФАСАДНЫЕ РАБОТЫ / Профиль стойка-ригель / Здание',
        'works': [{'name': 'Сборка и монтаж алюминиевых витражей', 'unit': 'м2'}],
        'materials': [
            {'name': 'Кронштейны опорные, ветровые', 'unit': 'шт', 'kind': 'вспомогат.', 'j': 1.5, 'k': 1, 'price': 760.22},
            {'name': 'Анкер клиновой 10*95', 'unit': 'шт', 'kind': 'вспомогат.', 'j': 3, 'k': 1.05, 'price': 69},
            {'name': 'Профиль алюминиевый (по проекту)', 'unit': 'м2', 'kind': 'основн.', 'j': 1, 'k': 1.1},
        ],
    },
    'spk_glass': {
        'costPath': 'ФАСАДНЫЕ РАБОТЫ / Светопрозрачные конструкции / Здание',
        'works': [{'name': 'Заполнение алюминиевых СПК', 'unit': 'м2'}],
        'materials': [{'name': 'Стеклопакет (по проекту)', 'unit': 'м2', 'kind': 'основн.', 'j': 0.9, 'k': 1.2}],
    },
    'spk_broneplenka': {
        'costPath': 'ФАСАДНЫЕ РАБОТЫ / Защита светопрозрачных конструкций / Здание',
        'works': [{'name': 'Оклейка бронирующей пленки с СПК', 'unit': 'м2'}],
        'materials': [{'name': 'Пленка защитная НТК ОПТИМА 90 мкм', 'unit': 'м2', 'kind': 'основн.', 'j': 1, 'k': 1.2}],
    },
    'scaffolding': {
        'costPath': 'ФАСАДНЫЕ РАБОТЫ / Леса и люльки / Здание',
        'works': [{'name': 'Монтаж/демонтаж лесов, подмостей', 'unit': 'м2'}],
        'materials': [{'name': 'Анкера для лесов, хомут-стяжка и пр', 'unit': 'м2', 'kind': 'вспомогат.', 'j': 1, 'k': 1.05, 'price': 800}],
    },
    'kmd_spk': {
        'costPath': 'ПРОЕКТНЫЕ РАБОТЫ / Разработка РД (включая КМД на фасады) и авторский надзор / Здание',
        'works': [{'name': 'Разработка КМ/КМД СПК', 'unit': 'м2'}],
        'materials': [],
    },
    'kmd_nvf': {
        'costPath': 'ПРОЕКТНЫЕ РАБОТЫ / Разработка РД (включая КМД на фасады) и авторский надзор / Здание',
        'works': [{'name': 'Разработка КМ/КМД НВФ', 'unit': 'м2'}],
        'materials': [],
    },
    'nvf_subsystem': {
        'costPath': 'ФАСАДНЫЕ РАБОТЫ / Подсистема НВФ + утеплитель / Здание',
        'works': [{'name': 'Монтаж подсистемы НВФ', 'unit': 'м2'}],
        'materials': [{'name': 'Подсистема НВФ (по проекту)', 'unit': 'м2', 'kind': 'основн.', 'j': 1, 'k': 1.12}],
    },
    'nvf_cladding_cassette': {
        'costPath': 'ФАСАДНЫЕ РАБОТЫ / Облицовка НВФ / Здание',
        'works': [{'name': 'Наружная облицовка фасада', 'unit': 'м2'}],
        'materials': [{'name': 'Облицовочный материал (по проекту)', 'unit': 'м2', 'kind': 'основн.', 'j': 1, 'k': 1.22}],
    },
    'insulation': {
        'costPath': 'ФАСАДНЫЕ РАБОТЫ / Подсистема НВФ + утеплитель / Здание',
        'works': [{'name': 'Утепление в 2 слоя (180 мм)', 'unit': 'м2'}],
        'materials': [
            {'name': 'Мембрана ветрозащитная', 'unit': 'м2', 'kind': 'основн.', 'j': 1, 'k': 1.15},
            {'name': 'Дюбель фасадный EJOT (наруж.)', 'unit': 'м2', 'kind': 'вспомогат.', 'j': 5, 'k': 1.05, 'price': 13.25},
            {'name': 'Утеплитель ТЕХНОВЕНТ ОПТИМА 50мм', 'unit': 'м3', 'kind': 'основн.', 'j': 0.08, 'k': 1.15},
            {'name': 'Утеплитель ТЕХНОВЕНТ Н 130мм', 'unit': 'м3', 'kind': 'основн.', 'j': 0.1, 'k': 1.15},
            {'name': 'Дюбель фасадный EJOT (внутр.)', 'unit': 'м2', 'kind': 'вспомогат.', 'j': 10, 'k': 1.05, 'price': 18.73},
        ],
    },
    'flashings': {
        'costPath': 'ФАСАДНЫЕ РАБОТЫ / Облицовка НВФ / Здание',
        'works': [{'name': 'Изготовление и монтаж оцинкованных элементов', 'unit': 'м.п.'}],
        'materials': [
            {'name': 'Лист 0,7 оц. с полимерным покрытием', 'unit': 'м2', 'kind': 'основн.', 'j': 0.3, 'k': 1.2},
            {'name': 'Заклепка вытяжная 4,0*8', 'unit': 'шт', 'kind': 'вспомогат.', 'j': 6, 'k': 1, 'price': 1.44},
            {'name': 'Дюбель-гвоздь 6*60', 'unit': 'шт', 'kind': 'вспомогат.', 'j': 6, 'k': 1, 'price': 1.6},
        ],
    },
    'glass_railing': {
        'costPath': 'ФАСАДНЫЕ РАБОТЫ / Ограждения, козырьки, маркизы / Здание',
        'works': [{'name': 'Монтаж стеклянных ограждений', 'unit': 'м2'}],
        'materials': [
            {'name': 'Профиль алюминиевый зажимной', 'unit': 'м', 'kind': 'основн.', 'j': 0.83, 'k': 1.1},
            {'name': 'Триплекс UltraClear', 'unit': 'м2', 'kind': 'основн.', 'j': 1, 'k': 1.2},
            {'name': 'Уплотнитель EPDM', 'unit': 'м', 'kind': 'вспомогат.', 'j': 1, 'k': 1.05, 'price': 50},
            {'name': 'Крышка декоративная', 'unit': 'м', 'kind': 'вспомогат.', 'j': 1, 'k': 1.05, 'price': 200},
            {'name': 'Клипса зажимная', 'unit': 'шт', 'kind': 'вспомогат.', 'j': 4, 'k': 1.05, 'price': 450},
            {'name': 'Химический анкер', 'unit': 'шт', 'kind': 'вспомогат.', 'j': 0.5, 'k': 1.05, 'price': 1690},
        ],
    },
    'glass_canopy': {
        'costPath': 'ФАСАДНЫЕ РАБОТЫ / Ограждения, козырьки, маркизы / Здание',
        'works': [{'name': 'Монтаж стеклянных козырьков', 'unit': 'м2'}],
        'materials': [{'name': 'Козырек из триплекса', 'unit': 'м2', 'kind': 'основн.', 'j': 1, 'k': 1.2}],
    },
    'doors_aluminum': {
        'costPath': 'ФАСАДНЫЕ РАБОТЫ / Двери наружные по фасаду (входные и БКФН, тамбурные двери) / Здание',
        'works': [{'name': 'Монтаж алюминиевых дверных блоков', 'unit': 'м2'}],
        'materials': [{'name': 'Профиль дверной (по проекту)', 'unit': 'м2', 'kind': 'основн.', 'j': 1, 'k': 1.1}],
    },
    'vent_grilles': {
        'costPath': 'ФАСАДНЫЕ РАБОТЫ / Облицовка НВФ / Здание',
        'works': [{'name': 'Установка решётки', 'unit': 'м2'}],
        'materials': [{'name': 'Вентиляционная решетка (по проекту)', 'unit': 'м2', 'kind': 'основн.', 'j': 1, 'k': 1.2}],
    },
    'wet_facade': {
        'costPath': 'ФАСАДНЫЕ РАБОТЫ / Устройство мокрого фасада / Здание',
        'works': [{'name': 'Устройство мокрого фасада (штукатурный слой)', 'unit': 'м2'}],
        'materials': [{'name': 'Штукатурная смесь (по проекту)', 'unit': 'м2', 'kind': 'основн.', 'j': 1, 'k': 1.15}],
    },
    'mockup': {
        'costPath': 'МОКАП / Фасадные работы / Здание',
        'works': [{'name': 'Устройство мокап фасада', 'unit': 'компл'}],
        'materials': [{'name': 'Материалы для мокап фасада', 'unit': 'компл', 'kind': 'основн.', 'j': 1, 'k': 1}],
    },
}

# ─── Правила матчинга (первое совпадение побеждает) ───────────────────
MATCH_RULES = [
    # Пропуск — не наш раздел / вспомогательные строки заказчика
    {'keywords': ['архитектурн.*освещен', 'освещен.*фасад'], 'templates': [], 'secondary': []},
    {'keywords': ['^прочие материал', 'прочие матери'], 'templates': [], 'secondary': []},
    {'keywords': ['краска атмосфер', 'краска фасадн'], 'templates': [], 'secondary': []},
    # МОКАП
    {'keywords': ['мокап', 'mock-up', 'mockup'], 'templates': ['mockup'], 'secondary': []},

    {'keywords': ['профил', 'каркас'], 'templates': ['spk_profile'], 'secondary': ['scaffolding', 'kmd_spk']},
    {'keywords': ['защит.*стекл', 'бронир', 'защит.*спк', 'защит.*светопрозр'], 'templates': ['spk_broneplenka'], 'secondary': []},
    # Стеклянные ограждения — ДО общего правила spk_glass
    {'keywords': ['ограждени.*стекл', 'стекл.*ограждени', 'ограждени.*кровл'], 'templates': ['glass_railing'], 'secondary': []},
    # СПК полный цикл
    {'keywords': ['стоечно-ригельн'], 'templates': ['spk_profile', 'spk_glass', 'spk_broneplenka'], 'secondary': ['scaffolding', 'kmd_spk']},
    {'keywords': ['стеклопакет', 'остеклен', 'заполнен'], 'templates': ['spk_glass'], 'secondary': []},
    {'keywords': ['двер'], 'templates': ['spk_profile', 'doors_aluminum', 'spk_glass'], 'secondary': ['scaffolding', 'kmd_spk']},
    {'keywords': ['вентрешетк', 'ламел', 'решетк', 'решётк'], 'templates': ['vent_grilles'], 'secondary': []},
    {'keywords': [r'лес[аоы]', 'люльк', 'подмост', 'подмащив', 'монтаж.*лесов', 'демонтаж.*лесов'], 'templates': ['scaffolding'], 'secondary': []},
    {'keywords': ['горизонтальн.*фасад', 'подшив'], 'templates': ['nvf_cladding_cassette'], 'secondary': []},
    {'keywords': ['подсистем'], 'templates': ['nvf_subsystem'], 'secondary': []},
    {'keywords': ['облицовк', 'облицовки'], 'templates': ['nvf_cladding_cassette'], 'secondary': ['scaffolding', 'kmd_nvf']},
    # "Утеплитель <название>" без глагола — строка-материал заказчика, пропускаем
    {'keywords': ['утеплитель толщ', 'утеплитель плита', 'утеплитель фасад баттс', 'утеплитель.*гост'], 'templates': [], 'secondary': []},
    {'keywords': ['утеплит', 'утеплени', 'утеплен.*фасад', 'минерал', 'минват', 'теплоизол'], 'templates': ['insulation'], 'secondary': []},
    {'keywords': ['откос'], 'templates': ['flashings'], 'secondary': []},
    {'keywords': ['отлив'], 'templates': ['flashings'], 'secondary': []},
    {'keywords': ['парапет'], 'templates': ['flashings'], 'secondary': []},
    {'keywords': ['козыр'], 'templates': ['glass_canopy'], 'secondary': []},
    {'keywords': ['штукатур', 'сфтк', 'мокр.*фасад', 'отделка стен'], 'templates': ['wet_facade'], 'secondary': []},
    # Составные (когда не разбиты)
    {'keywords': ['витраж', 'светопрозрачн'], 'templates': ['spk_profile', 'spk_glass'], 'secondary': ['scaffolding', 'kmd_spk']},
    {'keywords': ['окна', 'оконн'], 'templates': ['spk_profile', 'spk_glass'], 'secondary': ['scaffolding', 'kmd_spk']},
    {'keywords': ['непрозрачн'], 'templates': ['spk_profile', 'spk_glass'], 'secondary': ['scaffolding', 'kmd_spk']},
    {'keywords': ['навесн.*вентилир', 'вентилируем.*фасад'], 'templates': ['nvf_subsystem', 'nvf_cladding_cassette'], 'secondary': ['scaffolding', 'kmd_nvf']},
    {'keywords': ['кассет', 'сэндвич'], 'templates': ['nvf_subsystem', 'nvf_cladding_cassette'], 'secondary': ['scaffolding', 'kmd_nvf']},
    {'keywords': ['декоративн.*элемент'], 'templates': ['nvf_cladding_cassette'], 'secondary': []},
]


def is_header(pos, all_positions):
    """Позиция = заголовок, если есть дочерние с более длинным кодом."""
    if not pos['code']:
        return False
    prefix = pos['code'].rstrip('.')
    return any(
        o is not pos and o['code'] and
        o['code'].startswith(prefix + '.') and len(o['code']) > len(pos['code'])
        for o in all_positions
    )


def match_position(pos_name, note_customer=''):
    """Первое совпадение побеждает."""
    search_text = (pos_name + ' ' + (note_customer or '')).lower()
    matched = []
    seen = set()

    for rule in MATCH_RULES:
        hit = False
        for kw in rule['keywords']:
            if '.*' in kw or '[' in kw:
                if re.search(kw, search_text):
                    hit = True
                    break
            elif kw in search_text:
                hit = True
                break
        if hit:
            for t in rule['templates'] + rule['secondary']:
                if t not in seen:
                    matched.append(t)
                    seen.add(t)
            break  # первое совпадение побеждает

    return matched


def detect_vor_style(positions):
    """Определяет стиль ВОР: simple (Сокольники) или split-3 (Муза)."""
    aux_re = re.compile(r'прочие\s+материал|вспомогательн\w*\s+материал', re.I)
    return 'split-3' if any(aux_re.search(p['name'] or '') for p in positions) else 'simple'


def classify_row_role(name):
    """Классифицирует роль строки в split-3 режиме: work / material / auxiliary."""
    lower = (name or '').lower().strip()
    if re.search(r'прочие\s+материал|вспомогательн\w*\s+материал', lower):
        return 'auxiliary'
    if re.match(r'(монтаж|устройство|установка|сборка|демонтаж|оклейка|затирка|изготовление|разработка|утепление\s|наружная\s+облицовк|заполнение|монтаж/демонтаж)', lower):
        return 'work'
    return 'material'


def parse_empty_vor(filepath):
    """Парсим ВОР — все позиции плоским списком."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    print(f"Лист: {ws.title}, строк: {ws.max_row}, столбцов: {ws.max_column}")

    positions = []
    for row_idx in range(2, ws.max_row + 1):
        A = ws.cell(row_idx, 1).value
        E = ws.cell(row_idx, 5).value
        G = ws.cell(row_idx, 7).value
        H = ws.cell(row_idx, 8).value
        I_val = ws.cell(row_idx, 9).value
        S_val = ws.cell(row_idx, 19).value

        e_str = str(E or '').strip().lower()
        if 'суб-раб' in e_str or 'суб-мат' in e_str:
            continue

        g_str = str(G or '').strip()
        if not g_str or len(g_str) < 3:
            continue

        has_qty = I_val is not None and I_val != ''
        if not has_qty:
            continue

        pos_code = str(A or '').strip()
        qty_i = float(I_val) if isinstance(I_val, (int, float)) else 0

        positions.append({
            'row': row_idx,
            'code': pos_code,
            'name': g_str,
            'unit': str(H or '').strip(),
            'qty': qty_i,
            'note_customer': str(S_val or '').strip(),
        })

    wb.close()
    return positions


def generate_filled_vor(positions, output_path):
    """Генерируем заполненный Excel."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'ВОР расценённый'

    headers = ['Номер позиции', '№ п/п', 'Затрата на строительство', 'Наличие',
               'Тип элемента', 'Тип материала', 'Наименование', 'Ед. изм.',
               'Кол-во заказчика', 'Коэфф. перевода', 'Коэфф. расхода',
               'Кол-во ГП', 'Валюта', 'Тип доставки', 'Стоим. доставки',
               'Цена за единицу', 'Итоговая сумма', 'Ссылка на КП',
               'Примечание заказчика', 'Примечание ГП']

    for col, h in enumerate(headers, 1):
        cell = ws.cell(1, col, h)
        cell.fill = FILL_HEADER
        cell.font = Font(bold=True, size=10)
        cell.border = THIN_BORDER

    # Предварительный проход: определить excluded
    all_names = ' | '.join((p['name'] + ' ' + p.get('note_customer', '')).lower() for p in positions)
    exclude = set()
    if re.search(r'лес[аы]|люльк|подмост|подмащив', all_names):
        exclude.add('scaffolding')
    if re.search(r'защит.*стекл|бронир|защит.*спк', all_names):
        exclude.add('spk_broneplenka')

    row_num = 2
    total_matched = 0
    total_headers = 0
    total_works = 0
    total_materials = 0
    unmatched = []

    vor_style = detect_vor_style(positions)
    print(f"Стиль ВОР: {vor_style}")

    cluster_templates = []

    for pos in positions:
        # Позиция (розовая)
        ws.cell(row_num, 1, pos['code'])
        ws.cell(row_num, 7, pos['name'])
        ws.cell(row_num, 8, pos['unit'])
        ws.cell(row_num, 9, pos['qty'])
        ws.cell(row_num, 19, pos.get('note_customer', ''))
        for c in range(1, 21):
            ws.cell(row_num, c).fill = FILL_POSITION
            ws.cell(row_num, c).font = Font(bold=True, size=10)
            ws.cell(row_num, c).border = THIN_BORDER
        row_num += 1

        # Заголовки — пропускаем
        if is_header(pos, positions):
            total_headers += 1
            continue

        # ─── split-3 режим (Муза) ───────────────────────────────────
        if vor_style == 'split-3':
            role = classify_row_role(pos['name'])

            if role == 'auxiliary':
                # "Прочие материалы" — вспомогательные из кластера
                if not cluster_templates:
                    unmatched.append(f"{pos['code']:16s} {pos['name'][:55]}")
                    continue
                total_matched += 1

                # Пустая работа для привязки
                first_tpl = TEMPLATES.get(cluster_templates[0])
                if first_tpl and first_tpl['works']:
                    w = first_tpl['works'][0]
                    ws.cell(row_num, 3, first_tpl['costPath'])
                    ws.cell(row_num, 5, 'суб-раб')
                    ws.cell(row_num, 7, w['name'])
                    ws.cell(row_num, 8, w['unit'])
                    ws.cell(row_num, 13, 'RUB')
                    for c in range(1, 21):
                        ws.cell(row_num, c).fill = FILL_WORK
                        ws.cell(row_num, c).border = THIN_BORDER
                    total_works += 1
                    row_num += 1

                # Все вспомогательные материалы из кластера
                seen = set()
                for key in cluster_templates:
                    tpl = TEMPLATES.get(key)
                    if not tpl:
                        continue
                    for m in tpl['materials']:
                        if m.get('kind') != 'вспомогат.':
                            continue
                        if m['name'] in seen:
                            continue
                        seen.add(m['name'])
                        ws.cell(row_num, 3, tpl['costPath'])
                        ws.cell(row_num, 4, 'да')
                        ws.cell(row_num, 5, 'суб-мат')
                        ws.cell(row_num, 6, m.get('kind', 'основн.'))
                        ws.cell(row_num, 7, m['name'])
                        ws.cell(row_num, 8, m['unit'])
                        ws.cell(row_num, 10, m.get('j', 1))
                        ws.cell(row_num, 11, m.get('k', 1))
                        ws.cell(row_num, 13, 'RUB')
                        ws.cell(row_num, 14, 'в цене')
                        if m.get('price'):
                            ws.cell(row_num, 16, m['price'])
                        for c in range(1, 21):
                            ws.cell(row_num, c).fill = FILL_MATERIAL
                            ws.cell(row_num, c).border = THIN_BORDER
                        total_materials += 1
                        row_num += 1

                cluster_templates = []  # сброс
                continue

            # work или material — матчим
            raw_keys = match_position(pos['name'], pos.get('note_customer', ''))
            search_text = (pos['name'] + ' ' + pos.get('note_customer', '')).lower()
            matched_rule = None
            for rule in MATCH_RULES:
                hit = any(
                    (re.search(kw, search_text) if ('.*' in kw or '[' in kw) else kw in search_text)
                    for kw in rule['keywords']
                )
                if hit:
                    matched_rule = rule
                    break
            primary_set = set(matched_rule['templates']) if matched_rule else set()
            tpl_keys = [k for k in raw_keys if k not in exclude or k in primary_set]

            if matched_rule is not None and not matched_rule['templates'] and not matched_rule['secondary']:
                continue

            if not tpl_keys:
                unmatched.append(f"{pos['code']:16s} {pos['name'][:55]}")
                continue

            total_matched += 1

            # Накапливаем для auxiliary
            for k in tpl_keys:
                if k not in cluster_templates:
                    cluster_templates.append(k)

            if role == 'work':
                # ТОЛЬКО работы
                for key in tpl_keys:
                    tpl = TEMPLATES.get(key)
                    if not tpl:
                        continue
                    for w in tpl['works']:
                        ws.cell(row_num, 3, tpl['costPath'])
                        ws.cell(row_num, 5, 'суб-раб')
                        ws.cell(row_num, 7, w['name'])
                        ws.cell(row_num, 8, w['unit'])
                        ws.cell(row_num, 13, 'RUB')
                        for c in range(1, 21):
                            ws.cell(row_num, c).fill = FILL_WORK
                            ws.cell(row_num, c).border = THIN_BORDER
                        total_works += 1
                        row_num += 1
            else:
                # material: пустая работа + основные материалы
                for key in tpl_keys:
                    tpl = TEMPLATES.get(key)
                    if not tpl:
                        continue

                    if tpl['works']:
                        w = tpl['works'][0]
                        ws.cell(row_num, 3, tpl['costPath'])
                        ws.cell(row_num, 5, 'суб-раб')
                        ws.cell(row_num, 7, w['name'])
                        ws.cell(row_num, 8, w['unit'])
                        ws.cell(row_num, 13, 'RUB')
                        for c in range(1, 21):
                            ws.cell(row_num, c).fill = FILL_WORK
                            ws.cell(row_num, c).border = THIN_BORDER
                        total_works += 1
                        row_num += 1

                    for m in tpl['materials']:
                        if m.get('kind') != 'основн.':
                            continue
                        ws.cell(row_num, 3, tpl['costPath'])
                        ws.cell(row_num, 4, 'да')
                        ws.cell(row_num, 5, 'суб-мат')
                        ws.cell(row_num, 6, m.get('kind', 'основн.'))
                        ws.cell(row_num, 7, m['name'])
                        ws.cell(row_num, 8, m['unit'])
                        ws.cell(row_num, 10, m.get('j', 1))
                        ws.cell(row_num, 11, m.get('k', 1))
                        ws.cell(row_num, 13, 'RUB')
                        ws.cell(row_num, 14, 'в цене')
                        if m.get('price'):
                            ws.cell(row_num, 16, m['price'])
                        for c in range(1, 21):
                            ws.cell(row_num, c).fill = FILL_MATERIAL
                            ws.cell(row_num, c).border = THIN_BORDER
                        total_materials += 1
                        row_num += 1
            continue

        # ─── simple режим (Сокольники) ───────────────────────────────
        # Матчинг
        raw_keys = match_position(pos['name'], pos.get('note_customer', ''))
        matched_rule = None
        search_text = (pos['name'] + ' ' + pos.get('note_customer', '')).lower()
        for rule in MATCH_RULES:
            hit = any(
                (re.search(kw, search_text) if ('.*' in kw or '[' in kw) else kw in search_text)
                for kw in rule['keywords']
            )
            if hit:
                matched_rule = rule
                break
        primary_set = set(matched_rule['templates']) if matched_rule else set()
        tpl_keys = [k for k in raw_keys if k not in exclude or k in primary_set]

        # Намеренный пропуск: правило найдено, но templates=[] → пропускаем без ошибки
        if matched_rule is not None and not matched_rule['templates'] and not matched_rule['secondary']:
            continue

        if not tpl_keys:
            unmatched.append(f"{pos['code']:16s} {pos['name'][:55]}")
            continue

        total_matched += 1

        for key in tpl_keys:
            tpl = TEMPLATES.get(key)
            if not tpl:
                continue

            cost_path = tpl['costPath']

            for w in tpl['works']:
                ws.cell(row_num, 3, cost_path)
                ws.cell(row_num, 5, 'суб-раб')
                ws.cell(row_num, 7, w['name'])
                ws.cell(row_num, 8, w['unit'])
                ws.cell(row_num, 13, 'RUB')
                for c in range(1, 21):
                    ws.cell(row_num, c).fill = FILL_WORK
                    ws.cell(row_num, c).border = THIN_BORDER
                total_works += 1
                row_num += 1

            for m in tpl['materials']:
                ws.cell(row_num, 3, cost_path)
                ws.cell(row_num, 4, 'да')
                ws.cell(row_num, 5, 'суб-мат')
                ws.cell(row_num, 6, m.get('kind', 'основн.'))
                ws.cell(row_num, 7, m['name'])
                ws.cell(row_num, 8, m['unit'])
                ws.cell(row_num, 10, m.get('j', 1))
                ws.cell(row_num, 11, m.get('k', 1))
                ws.cell(row_num, 13, 'RUB')
                ws.cell(row_num, 14, 'в цене')
                if m.get('price'):
                    ws.cell(row_num, 16, m['price'])
                for c in range(1, 21):
                    ws.cell(row_num, c).fill = FILL_MATERIAL
                    ws.cell(row_num, c).border = THIN_BORDER
                total_materials += 1
                row_num += 1

    col_widths = [14, 6, 40, 8, 10, 12, 55, 8, 14, 12, 12, 14, 8, 12, 12, 14, 16, 20, 20, 20]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    wb.save(output_path)
    print(f"\n{'='*60}")
    print(f"РЕЗУЛЬТАТ: {output_path}")
    print(f"  Стиль: {vor_style}")
    print(f"  Позиций: {len(positions)}")
    print(f"  Заголовков: {total_headers}")
    print(f"  Заматчено: {total_matched}")
    print(f"  Работ: {total_works}")
    print(f"  Материалов: {total_materials}")
    if unmatched:
        print(f"\n  Не заматчено ({len(unmatched)}):")
        for u in unmatched:
            print(f"    - {u}")
    print(f"{'='*60}")


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("Usage: python fill_vor.py input.xlsx [output.xlsx]")
        sys.exit(1)

    input_path = sys.argv[1]
    output_path = sys.argv[2] if len(sys.argv) > 2 else input_path.replace('.xlsx', '_filled.xlsx')

    positions = parse_empty_vor(input_path)
    print(f"Всего позиций: {len(positions)}")
    generate_filled_vor(positions, output_path)
