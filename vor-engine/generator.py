"""Excel workbook generator for VOR pipeline results.

Produces a 4-sheet ``.xlsx`` workbook:

* **ВОР** — Main bill of quantities with all line items.
* **AI Комментарии** — Notes, warnings, and issues raised during processing.
* **Сводка** — Summary with total cost, section breakdown, and confidence stats.
* **Ресурсная ведомость** — Resource breakdown (labor, machinery, materials)
  from GESN norms, color-coded by resource type.
"""

from __future__ import annotations

import io
from pathlib import Path
from typing import Any

import yaml
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Market price ranges (loaded once from price_ranges.yaml)
# ---------------------------------------------------------------------------

def _load_market_ranges() -> dict[str, dict[str, Any]]:
    """Load position_totals section from price_ranges.yaml."""
    yaml_path = Path(__file__).parent / "price_ranges.yaml"
    if not yaml_path.exists():
        return {}
    try:
        data = yaml.safe_load(yaml_path.read_text(encoding="utf-8"))
        return data.get("position_totals", {})
    except Exception:
        return {}


_MARKET_RANGES: dict[str, dict[str, Any]] = _load_market_ranges()


def _lookup_market_range(item_name: str, item_unit: str) -> tuple[str, bool]:
    """Find market price range for an item by keyword matching.

    Returns (display_string, is_in_range) where display_string is
    "min–max" or "—" if no match. is_in_range is always True when
    no range is found (no flag needed).
    """
    name_lower = item_name.lower()
    for _key, info in _MARKET_RANGES.items():
        keywords = info.get("keywords", [])
        if any(kw in name_lower for kw in keywords):
            return f"{info['min']:,.0f}–{info['max']:,.0f}", True
    return "—", True

from vor.models import PositionBreakdown, ResourceDetail, ResourceLine, VorResult, WorkBreakdown

# ---------------------------------------------------------------------------
# Style constants
# ---------------------------------------------------------------------------

_FILL_GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
_FILL_YELLOW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
_FILL_RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
_FILL_HEADER = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_FILL_SECTION = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")

_FONT_HEADER = Font(bold=True, color="FFFFFF", size=11)
_FONT_BOLD = Font(bold=True, size=11)
_FONT_NORMAL = Font(size=10)
_FONT_WARNING = Font(bold=True, color="FF0000", size=11)

_ALIGN_WRAP = Alignment(wrap_text=True, vertical="top")
_ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_ALIGN_RIGHT = Alignment(horizontal="right", vertical="top")

_THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

_CONFIDENCE_FILLS = {
    "green": _FILL_GREEN,
    "yellow": _FILL_YELLOW,
    "red": _FILL_RED,
}

# Resource-type fills for the resource breakdown sheet
_FILL_RES_LABOR = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")       # light blue
_FILL_RES_MACHINERY = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")   # light yellow
_FILL_RES_MATERIAL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")    # light green

_RESOURCE_TYPE_FILLS = {
    "labor": _FILL_RES_LABOR,
    "labor_operator": _FILL_RES_LABOR,
    "machinery": _FILL_RES_MACHINERY,
    "material": _FILL_RES_MATERIAL,
}

# ---------------------------------------------------------------------------
# Column definitions for Sheet 1 (ВОР)
# ---------------------------------------------------------------------------

_VOR_COLUMNS = [
    ("№", 6),
    ("Раздел", 25),
    ("Наименование", 50),
    ("Ед.изм.", 10),
    ("Кол-во", 12),
    ("Шифр ГЭСН", 18),
    ("Расценка (база 2000)", 18),
    ("Стоимость (база 2000)", 20),
    ("Уверенность AI", 16),
    ("Причина", 40),
    ("Рынок 2025, ₽/ед", 20),
    ("Примечание", 35),
]

# Column definitions for Sheet 2 (AI Комментарии)
_COMMENT_COLUMNS = [
    ("№", 6),
    ("Позиция", 8),
    ("Уровень", 12),
    ("Комментарий", 80),
]

# Column definitions for Sheet 3 (Сводка)
_SUMMARY_COLUMNS = [
    ("Показатель", 40),
    ("Значение", 25),
]

# Column definitions for Sheet 4 (Ресурсная ведомость)
_RESOURCE_COLUMNS = [
    ("№ п/п", 8),
    ("Код ГЭСН", 18),
    ("Наименование работы", 40),
    ("Код ресурса", 18),
    ("Наименование ресурса", 50),
    ("Тип", 16),
    ("Ед.изм.", 12),
    ("Норма", 14),
    ("Кол-во", 14),
]


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------


def generate_vor_excel_from_dicts(data: dict) -> bytes:
    """Generate Excel from dict-serialized VOR result (loaded from disk).

    Reconstructs VorResult from plain dicts, then delegates to generate_vor_excel.
    """
    from vor.models import VorItem, GesnMatch, PriceResult, QuantityPlan

    items = [VorItem(**{k: v for k, v in d.items() if k in ('row_num', 'name', 'unit', 'quantity', 'section', 'raw_data')})
             for d in data.get("items", [])]
    matches = [GesnMatch(**{k: v for k, v in d.items() if k in ('item_idx', 'gesn_code', 'gesn_name', 'gesn_unit', 'confidence', 'confidence_level', 'alternatives', 'reasoning')})
               for d in data.get("matches", [])]
    prices = [PriceResult(**{k: v for k, v in d.items() if k in ('item_idx', 'quantity', 'gesn_code', 'fer_direct_cost', 'fer_labor', 'fer_machinery', 'fer_materials', 'total_base', 'notes', 'resources')})
              for d in data.get("prices", [])]
    plans = [QuantityPlan(**{k: v for k, v in d.items() if k in ('item_idx', 'source', 'category', 'parameter', 'filter_criteria', 'unit_conversion', 'notes')})
             for d in data.get("plans", [])]

    result = VorResult(
        items=items, matches=matches, prices=prices, plans=plans,
        stats=data.get("stats", {}), errors=data.get("errors", []),
    )
    return generate_vor_excel(result)


def generate_vor_excel(result: VorResult) -> bytes:
    """Generate a 4-sheet Excel workbook from VOR pipeline results.

    Returns the workbook as raw bytes suitable for writing to a file or
    sending over HTTP.
    """
    wb = Workbook()

    # Sheet 1: ВОР (main bill of quantities)
    ws_vor = wb.active
    ws_vor.title = "ВОР"
    _build_vor_sheet(ws_vor, result)

    # Sheet 2: AI Комментарии
    ws_comments = wb.create_sheet("AI Комментарии")
    _build_comments_sheet(ws_comments, result)

    # Sheet 3: Сводка
    ws_summary = wb.create_sheet("Сводка")
    _build_summary_sheet(ws_summary, result)

    # Sheet 4: Ресурсная ведомость
    ws_resources = wb.create_sheet("Ресурсная ведомость")
    _build_resource_sheet(ws_resources, result)

    # Write to bytes
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------


def _build_vor_sheet(ws: Any, result: VorResult) -> None:
    """Build the main ВОР sheet."""
    # Warning row
    ws.merge_cells("A1:L1")
    cell = ws["A1"]
    cell.value = "Цены в базе 2000 года. Для пересчёта в текущие цены применяйте индексы Минстроя."
    cell.font = _FONT_WARNING
    cell.alignment = _ALIGN_WRAP

    # Header row (row 2)
    header_row = 2
    for col_idx, (title, width) in enumerate(_VOR_COLUMNS, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=title)
        cell.font = _FONT_HEADER
        cell.fill = _FILL_HEADER
        cell.alignment = _ALIGN_CENTER
        cell.border = _THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Freeze header
    ws.freeze_panes = "A3"

    # Auto-filter
    ws.auto_filter.ref = f"A{header_row}:J{header_row}"

    # Build lookup dicts
    match_by_idx = {m.item_idx: m for m in result.matches}
    price_by_idx = {p.item_idx: p for p in result.prices}
    plan_by_idx = {p.item_idx: p for p in result.plans}

    # Data rows
    current_section = ""
    data_start_row = header_row + 1
    row_num = data_start_row

    for idx, item in enumerate(result.items):
        # Section header row
        if item.section and item.section != current_section:
            current_section = item.section
            ws.merge_cells(f"A{row_num}:L{row_num}")
            cell = ws.cell(row=row_num, column=1, value=current_section)
            cell.font = _FONT_BOLD
            cell.fill = _FILL_SECTION
            cell.alignment = _ALIGN_WRAP
            row_num += 1

        match = match_by_idx.get(idx)
        price = price_by_idx.get(idx)
        plan = plan_by_idx.get(idx)

        quantity = price.quantity if price else item.quantity
        gesn_code = match.gesn_code if match else ""
        fer_cost = price.fer_direct_cost if price else 0.0
        total_base = price.total_base if price else 0.0
        confidence_level = match.confidence_level if match else ""
        confidence_value = match.confidence if match else 0.0

        # Notes
        notes_parts: list[str] = []
        if match and match.reasoning:
            notes_parts.append(match.reasoning)
        if price and price.notes:
            notes_parts.append(price.notes)
        if plan and plan.notes:
            notes_parts.append(plan.notes)
        notes = "; ".join(notes_parts)

        # Reason text based on confidence level
        if confidence_level == "green":
            reason = "Точное совпадение кода"
        elif confidence_level == "yellow":
            reason = match.reasoning if match and match.reasoning else "Средняя уверенность"
        elif confidence_level == "red":
            reason = (match.reasoning if match and match.reasoning else "Низкая уверенность") + " — требуется проверка"
        else:
            reason = ""

        # Market price range lookup
        market_range, _in_range = _lookup_market_range(item.name, item.unit)

        # Row number (1-based for display)
        row_data = [
            idx + 1,                           # №
            item.section,                      # Раздел
            item.name,                         # Наименование
            item.unit,                         # Ед.изм.
            quantity,                           # Кол-во
            gesn_code,                         # Шифр ГЭСН
            fer_cost,                          # Расценка
            total_base,                        # Стоимость
            f"{confidence_value:.0%}",         # Уверенность AI
            reason,                            # Причина
            market_range,                      # Рынок 2025
            notes,                             # Примечание
        ]

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_num, column=col_idx, value=value)
            cell.font = _FONT_NORMAL
            cell.alignment = _ALIGN_WRAP
            cell.border = _THIN_BORDER

            # Number formatting
            if col_idx in (7, 8):  # Cost columns
                cell.number_format = "#,##0.00"
            elif col_idx == 5:  # Quantity
                cell.number_format = "#,##0.000"

        # Confidence coloring
        if confidence_level in _CONFIDENCE_FILLS:
            ws.cell(row=row_num, column=9).fill = _CONFIDENCE_FILLS[confidence_level]

        row_num += 1

    # Totals row
    if result.items:
        total_row = row_num
        ws.cell(row=total_row, column=1, value="").border = _THIN_BORDER
        ws.cell(row=total_row, column=2, value="ИТОГО").font = _FONT_BOLD
        ws.cell(row=total_row, column=2).border = _THIN_BORDER

        for col_idx in range(3, 11):
            ws.cell(row=total_row, column=col_idx).border = _THIN_BORDER

        # SUM formula for cost column (column 8)
        sum_cell = ws.cell(
            row=total_row,
            column=8,
            value=f"=SUM(H{data_start_row}:H{total_row - 1})",
        )
        sum_cell.font = _FONT_BOLD
        sum_cell.number_format = "#,##0.00"
        sum_cell.border = _THIN_BORDER


def _build_comments_sheet(ws: Any, result: VorResult) -> None:
    """Build the AI comments sheet."""
    # Header row
    for col_idx, (title, width) in enumerate(_COMMENT_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.font = _FONT_HEADER
        cell.fill = _FILL_HEADER
        cell.alignment = _ALIGN_CENTER
        cell.border = _THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"

    # Collect comments from various sources
    comment_num = 0
    row = 2

    # Comments from errors
    for error in result.errors:
        comment_num += 1
        position = error.get("item_idx", "")
        if isinstance(position, int):
            position = position + 1  # 1-based display
        ws.cell(row=row, column=1, value=comment_num).border = _THIN_BORDER
        ws.cell(row=row, column=2, value=position).border = _THIN_BORDER
        level_cell = ws.cell(row=row, column=3, value=error.get("stage", "error"))
        level_cell.border = _THIN_BORDER
        level_cell.fill = _FILL_RED
        ws.cell(row=row, column=4, value=error.get("message", "")).border = _THIN_BORDER
        row += 1

    # Comments from low-confidence matches
    for match in result.matches:
        if match.confidence_level == "red":
            comment_num += 1
            ws.cell(row=row, column=1, value=comment_num).border = _THIN_BORDER
            ws.cell(row=row, column=2, value=match.item_idx + 1).border = _THIN_BORDER
            level_cell = ws.cell(row=row, column=3, value="low confidence")
            level_cell.border = _THIN_BORDER
            level_cell.fill = _FILL_RED
            msg = f"Low confidence ({match.confidence:.0%}) for GESN match. {match.reasoning}"
            ws.cell(row=row, column=4, value=msg).border = _THIN_BORDER
            row += 1
        elif match.confidence_level == "yellow":
            comment_num += 1
            ws.cell(row=row, column=1, value=comment_num).border = _THIN_BORDER
            ws.cell(row=row, column=2, value=match.item_idx + 1).border = _THIN_BORDER
            level_cell = ws.cell(row=row, column=3, value="medium confidence")
            level_cell.border = _THIN_BORDER
            level_cell.fill = _FILL_YELLOW
            msg = f"Medium confidence ({match.confidence:.0%}) for GESN match. {match.reasoning}"
            ws.cell(row=row, column=4, value=msg).border = _THIN_BORDER
            row += 1

    # Comments from manual plans
    for plan in result.plans:
        if plan.source == "manual":
            comment_num += 1
            ws.cell(row=row, column=1, value=comment_num).border = _THIN_BORDER
            ws.cell(row=row, column=2, value=plan.item_idx + 1).border = _THIN_BORDER
            level_cell = ws.cell(row=row, column=3, value="manual input")
            level_cell.border = _THIN_BORDER
            level_cell.fill = _FILL_YELLOW
            msg = f"Quantity requires manual input. {plan.notes}"
            ws.cell(row=row, column=4, value=msg).border = _THIN_BORDER
            row += 1


def _build_summary_sheet(ws: Any, result: VorResult) -> None:
    """Build the summary sheet."""
    # Header
    for col_idx, (title, width) in enumerate(_SUMMARY_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.font = _FONT_HEADER
        cell.fill = _FILL_HEADER
        cell.alignment = _ALIGN_CENTER
        cell.border = _THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"

    # Compute stats
    total_items = len(result.items)
    green_count = sum(1 for m in result.matches if m.confidence_level == "green")
    yellow_count = sum(1 for m in result.matches if m.confidence_level == "yellow")
    red_count = sum(1 for m in result.matches if m.confidence_level == "red")
    not_found = sum(1 for m in result.matches if not m.gesn_code)
    total_cost = sum(p.total_base for p in result.prices)

    # Section breakdown
    section_costs: dict[str, float] = {}
    price_by_idx = {p.item_idx: p for p in result.prices}
    for idx, item in enumerate(result.items):
        section = item.section or "Без раздела"
        price = price_by_idx.get(idx)
        cost = price.total_base if price else 0.0
        section_costs[section] = section_costs.get(section, 0.0) + cost

    # Write summary rows
    rows = [
        ("Общие показатели", ""),
        ("Всего позиций", total_items),
        ("Общая стоимость (база 2000)", round(total_cost, 2)),
        ("", ""),
        ("Качество сопоставления", ""),
        ("Зелёный (высокая уверенность)", green_count),
        ("Жёлтый (средняя уверенность)", yellow_count),
        ("Красный (низкая уверенность)", red_count),
        ("Не найдено", not_found),
        ("", ""),
        ("Стоимость по разделам", ""),
    ]

    for section, cost in section_costs.items():
        rows.append((section, round(cost, 2)))

    for row_idx, (label, value) in enumerate(rows, start=2):
        label_cell = ws.cell(row=row_idx, column=1, value=label)
        value_cell = ws.cell(row=row_idx, column=2, value=value)
        label_cell.border = _THIN_BORDER
        value_cell.border = _THIN_BORDER

        # Bold for section headers (rows with empty value column that are not blank)
        if label and value == "":
            label_cell.font = _FONT_BOLD
            label_cell.fill = _FILL_SECTION
        else:
            label_cell.font = _FONT_NORMAL

        # Number formatting for cost values
        if isinstance(value, (int, float)) and value != 0:
            value_cell.number_format = "#,##0.00"

        # Color stats rows
        if label == "Зелёный (высокая уверенность)":
            value_cell.fill = _FILL_GREEN
        elif label == "Жёлтый (средняя уверенность)":
            value_cell.fill = _FILL_YELLOW
        elif label == "Красный (низкая уверенность)":
            value_cell.fill = _FILL_RED


# ---------------------------------------------------------------------------
# Resource type label mapping (for display)
# ---------------------------------------------------------------------------

_RESOURCE_TYPE_LABELS = {
    "labor": "Труд",
    "labor_operator": "Труд машиниста",
    "machinery": "Механизмы",
    "material": "Материалы",
}


def _build_resource_sheet(ws: Any, result: VorResult) -> None:
    """Build the resource breakdown sheet.

    Groups resources by VOR item (PriceResult). Each group has:
    - A section-header row with the GESN code and work name
    - One row per resource, color-coded by type
    """
    # Header row
    for col_idx, (title, width) in enumerate(_RESOURCE_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.font = _FONT_HEADER
        cell.fill = _FILL_HEADER
        cell.alignment = _ALIGN_CENTER
        cell.border = _THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"

    # Build helpers
    item_by_idx = {idx: item for idx, item in enumerate(result.items)}
    match_by_idx = {m.item_idx: m for m in result.matches}

    row_num = 2
    seq = 0  # sequential counter

    for price in result.prices:
        if not price.resources:
            continue

        item = item_by_idx.get(price.item_idx)
        item_name = item.name if item else ""
        match = match_by_idx.get(price.item_idx)
        gesn_name = match.gesn_name if match else ""
        work_label = item_name or gesn_name

        # Group header row (merged across all columns)
        num_cols = len(_RESOURCE_COLUMNS)
        ws.merge_cells(
            start_row=row_num, start_column=1,
            end_row=row_num, end_column=num_cols,
        )
        header_cell = ws.cell(
            row=row_num, column=1,
            value=f"{price.gesn_code}  —  {work_label}  (кол-во: {price.quantity})",
        )
        header_cell.font = _FONT_BOLD
        header_cell.fill = _FILL_SECTION
        header_cell.alignment = _ALIGN_WRAP
        header_cell.border = _THIN_BORDER
        row_num += 1

        # Resource rows
        for res in price.resources:
            seq += 1
            type_label = _RESOURCE_TYPE_LABELS.get(res.type, res.type)
            row_data = [
                seq,                  # № п/п
                price.gesn_code,      # Код ГЭСН
                work_label,           # Наименование работы
                res.resource_code,    # Код ресурса
                res.name,             # Наименование ресурса
                type_label,           # Тип
                res.measure_unit,     # Ед.изм.
                res.norm_quantity,    # Норма
                res.total_quantity,   # Кол-во
            ]

            fill = _RESOURCE_TYPE_FILLS.get(res.type)

            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_num, column=col_idx, value=value)
                cell.font = _FONT_NORMAL
                cell.alignment = _ALIGN_WRAP
                cell.border = _THIN_BORDER
                if fill:
                    cell.fill = fill
                # Number formatting for norm and quantity columns
                if col_idx in (8, 9):
                    cell.number_format = "#,##0.000000"

            row_num += 1


# ===========================================================================
# V3: Single-sheet generator with resource breakdown
# ===========================================================================

# Style constants for v3
_FILL_POSITION = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")  # blue-ish
_FILL_WORK = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")       # green-ish
_FILL_MAT_MAIN = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")   # warm yellow
_FILL_MAT_AUX = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")    # white
_FILL_MACHINERY_V3 = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")  # orange
_FILL_SUPPLEMENT = PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid")  # orange-ish
_FILL_COMMENT = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")    # light grey
_FILL_SECTION_TOTAL = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")

_FONT_POSITION = Font(bold=True, size=11)
_FONT_WORK = Font(bold=True, size=10, color="2E7D32")
_FONT_RESOURCE = Font(size=9)
_FONT_RESOURCE_MAIN = Font(size=9, bold=True)
_FONT_COMMENT = Font(size=9, italic=True, color="555555")
_FONT_SUPPLEMENT = Font(bold=True, size=10, color="C65102")

# Column definitions for v3
_V3_COLUMNS = [
    ("№", 5),
    ("Тип", 5),
    ("Наименование", 50),
    ("Шифр / Код", 18),
    ("Ед.изм.", 10),
    ("Кол-во", 12),
    ("Цена", 14),
    ("Стоимость", 16),
    ("Примечание", 40),
]


def generate_vor_excel_v3(result: VorResult) -> bytes:
    """Generate a single-sheet Excel workbook with resource breakdown.

    Structure per section:
      Section header
        Position 1 (total cost)
          Work (GESN code)
            ★ Main material (with price)
            ○ Aux material (with price)
          Work 2 ...
          ☰ Machinery
          💬 Comment
        Position 2 ...
        ДОПНИК: supplement items
        Итого по разделу
      Grand total
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "ВОР с ресурсной раскладкой"

    # Warning row
    num_cols = len(_V3_COLUMNS)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    cell = ws["A1"]
    cell.value = "Цены в базе 2000 года (ФССЦ). Для пересчёта в текущие цены применяйте индексы Минстроя."
    cell.font = _FONT_WARNING
    cell.alignment = _ALIGN_WRAP

    # Header row (row 2)
    for col_idx, (title, width) in enumerate(_V3_COLUMNS, start=1):
        cell = ws.cell(row=2, column=col_idx, value=title)
        cell.font = _FONT_HEADER
        cell.fill = _FILL_HEADER
        cell.alignment = _ALIGN_CENTER
        cell.border = _THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A3"

    row = 3

    # Group breakdowns by section
    breakdowns = result.breakdowns
    if not breakdowns:
        # Fallback: empty workbook
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.read()

    # Build section groups preserving order
    sections: list[tuple[str, list[PositionBreakdown]]] = []
    seen_sections: dict[str, int] = {}

    # Regular positions
    item_by_idx = {idx: item for idx, item in enumerate(result.items)}
    for bd in breakdowns:
        if bd.is_supplement:
            continue
        item = item_by_idx.get(bd.item_idx)
        sec = item.section if item else "Без раздела"
        if sec not in seen_sections:
            seen_sections[sec] = len(sections)
            sections.append((sec, []))
        sections[seen_sections[sec]][1].append(bd)

    # Supplements: attach to the section they logically belong to, or last section
    supplements = [bd for bd in breakdowns if bd.is_supplement]

    grand_total = 0.0
    pos_num = 0

    for sec_name, sec_breakdowns in sections:
        # Section header
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)
        sec_cell = ws.cell(row=row, column=1, value=f"РАЗДЕЛ: {sec_name}")
        sec_cell.font = _FONT_BOLD
        sec_cell.fill = _FILL_SECTION
        sec_cell.alignment = _ALIGN_WRAP
        sec_cell.border = _THIN_BORDER
        row += 1

        section_total = 0.0

        for bd in sec_breakdowns:
            pos_num += 1
            row = _write_position(ws, row, pos_num, bd, num_cols)
            section_total += bd.total_cost

        # Supplements for this section (attach all to last section for now)
        if sec_name == sections[-1][0] and supplements:
            # Supplement header
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)
            sup_header = ws.cell(row=row, column=1, value="ДОПНИКИ (дополнительные позиции)")
            sup_header.font = _FONT_SUPPLEMENT
            sup_header.fill = _FILL_SUPPLEMENT
            sup_header.border = _THIN_BORDER
            row += 1

            for sup_idx, sup in enumerate(supplements, start=1):
                row = _write_position(ws, row, f"Д{sup_idx}", sup, num_cols, is_supplement=True)
                section_total += sup.total_cost

        # Section total
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
        total_label = ws.cell(row=row, column=1, value=f"ИТОГО по разделу «{sec_name}»:")
        total_label.font = _FONT_BOLD
        total_label.fill = _FILL_SECTION_TOTAL
        total_label.border = _THIN_BORDER
        total_cell = ws.cell(row=row, column=8, value=round(section_total, 2))
        total_cell.font = _FONT_BOLD
        total_cell.fill = _FILL_SECTION_TOTAL
        total_cell.number_format = "#,##0.00"
        total_cell.border = _THIN_BORDER
        ws.cell(row=row, column=9).border = _THIN_BORDER
        row += 1
        row += 1  # blank row between sections

        grand_total += section_total

    # Grand total
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    gt_label = ws.cell(row=row, column=1, value="ИТОГО ПО ВОР:")
    gt_label.font = Font(bold=True, size=12, color="FFFFFF")
    gt_label.fill = _FILL_HEADER
    gt_label.border = _THIN_BORDER
    gt_cell = ws.cell(row=row, column=8, value=round(grand_total, 2))
    gt_cell.font = Font(bold=True, size=12, color="FFFFFF")
    gt_cell.fill = _FILL_HEADER
    gt_cell.number_format = "#,##0.00"
    gt_cell.border = _THIN_BORDER
    ws.cell(row=row, column=9).border = _THIN_BORDER
    ws.cell(row=row, column=9).fill = _FILL_HEADER

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _write_position(
    ws: Any,
    row: int,
    pos_num: int | str,
    bd: PositionBreakdown,
    num_cols: int,
    *,
    is_supplement: bool = False,
) -> int:
    """Write a single position with its works and resources. Returns next row."""
    # Position header row
    conf_map = {"green": _FILL_GREEN, "yellow": _FILL_YELLOW, "red": _FILL_RED}
    pos_fill = _FILL_SUPPLEMENT if is_supplement else _FILL_POSITION

    pos_data = [
        pos_num,                         # №
        "",                              # Тип
        bd.item_name,                    # Наименование
        "",                              # Шифр
        bd.unit,                         # Ед.изм.
        bd.quantity if bd.quantity else "",  # Кол-во
        "",                              # Цена (n/a for position)
        bd.total_cost,                   # Стоимость
        "",                              # Примечание
    ]

    for col_idx, value in enumerate(pos_data, start=1):
        cell = ws.cell(row=row, column=col_idx, value=value)
        cell.font = _FONT_POSITION
        cell.fill = pos_fill
        cell.alignment = _ALIGN_WRAP
        cell.border = _THIN_BORDER
        if col_idx == 8:
            cell.number_format = "#,##0.00"

    # Confidence fill on the note column
    if bd.confidence_level in conf_map:
        ws.cell(row=row, column=9).fill = conf_map[bd.confidence_level]
        ws.cell(row=row, column=9).value = f"{bd.confidence:.0%}"

    row += 1

    # Works and their resources
    for wb_item in bd.works:
        # Work line
        work_data = [
            "",                                 # №
            "Р",                                # Тип (Работа)
            wb_item.gesn_name or wb_item.gesn_code,  # Наименование
            wb_item.gesn_code,                  # Шифр
            wb_item.measure_unit,               # Ед.изм.
            wb_item.quantity,                    # Кол-во
            "",                                 # Цена
            wb_item.total_cost,                 # Стоимость
            wb_item.reasoning[:80] if wb_item.reasoning else "",  # Примечание
        ]

        for col_idx, value in enumerate(work_data, start=1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.font = _FONT_WORK
            cell.fill = _FILL_WORK
            cell.alignment = _ALIGN_WRAP
            cell.border = _THIN_BORDER
            if col_idx == 8:
                cell.number_format = "#,##0.00"

        row += 1

        # Materials (main first, then auxiliary)
        sorted_materials = sorted(wb_item.materials, key=lambda m: (not m.is_main, m.name))
        for mat in sorted_materials:
            row = _write_resource_line(ws, row, mat)

        # Machinery
        for mach in wb_item.machinery:
            row = _write_resource_line(ws, row, mach, is_machinery=True)

    # Comment row (if any)
    if bd.comment:
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=num_cols)
        ws.cell(row=row, column=1).border = _THIN_BORDER
        comment_cell = ws.cell(row=row, column=2, value=f"AI: {bd.comment}")
        comment_cell.font = _FONT_COMMENT
        comment_cell.fill = _FILL_COMMENT
        comment_cell.alignment = _ALIGN_WRAP
        comment_cell.border = _THIN_BORDER
        row += 1

    return row


def _write_resource_line(
    ws: Any,
    row: int,
    rl: ResourceLine,
    *,
    is_machinery: bool = False,
) -> int:
    """Write a single resource line. Returns next row."""
    if is_machinery:
        prefix = chr(9776)  # ☰
        fill = _FILL_MACHINERY_V3
        font = _FONT_RESOURCE
        type_code = "М"  # Механизм
    elif rl.is_main:
        prefix = chr(9733)  # ★
        fill = _FILL_MAT_MAIN
        font = _FONT_RESOURCE_MAIN
        type_code = "МО"  # Материал основной
    elif rl.resource_type == "material_unaccounted":
        prefix = chr(9675)  # ○
        fill = _FILL_MAT_AUX
        font = _FONT_RESOURCE
        type_code = "МН"  # Материал неучтённый
    else:
        prefix = chr(9675)  # ○
        fill = _FILL_MAT_AUX
        font = _FONT_RESOURCE
        type_code = "МВ"  # Материал вспомогательный

    price_note = "" if rl.price_found else "цена не найдена"
    note = rl.note or price_note

    res_data = [
        "",                              # №
        type_code,                       # Тип
        f"  {prefix} {rl.name}",         # Наименование (indented)
        rl.resource_code,                # Код ресурса
        rl.measure_unit,                 # Ед.изм.
        rl.total_quantity,               # Кол-во
        rl.unit_price if rl.price_found else 0,  # Цена
        rl.total_price,                  # Стоимость
        note,                            # Примечание
    ]

    for col_idx, value in enumerate(res_data, start=1):
        cell = ws.cell(row=row, column=col_idx, value=value)
        cell.font = font
        cell.fill = fill
        cell.alignment = _ALIGN_WRAP
        cell.border = _THIN_BORDER
        if col_idx in (7, 8):
            cell.number_format = "#,##0.00"
        elif col_idx == 6:
            cell.number_format = "#,##0.000"

    return row + 1
